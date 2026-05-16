"""Convert IBGE PNADC Cultural Sector tables (Tabela 6.1a–6.17) to Parquet.

Strategy: each xlsx → one Parquet, wide format, preserves IBGE's column structure.
- One row per (year, row_label)
- row_label = first cell value (region name, position label, etc.)
- columns c01, c02, ... = original column values (incl. CV in c02, c04, c06, ...)

For column meaning, refer to docs/column_maps_pnadc.md (mirror of
.claude/skills/ibge-pnadc-cultural/references/column_maps.md).

Tables 6.1a and 6.1b have geography in COLUMNS not rows — handled as special case.

Usage:
    export MOTHERDUCK_TOKEN="..."
    python etl/ibge_pnadc__xlsx_to_parquet.py
"""
import os
import sys
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "raw" / "ibge_pnadc"
OUT.mkdir(parents=True, exist_ok=True)

SRC = Path(
    "/Users/joaoroque/Documents/Cultural production - book/"
    "Dados da Economia Cultural no Brasil/"
    "6_ocupacao_no_setor_cultural"
)
if not SRC.exists():
    SRC = Path(
        "/sessions/wizardly-admiring-feynman/mnt/"
        "Dados da Economia Cultural no Brasil/"
        "6_ocupacao_no_setor_cultural"
    )

YEARS = [str(y) for y in range(2014, 2025)]

# Tables with geographic ROWS (standard case)
GEO_ROW_TABLES = ["6.3", "6.4", "6.5", "6.6", "6.7", "6.8", "6.10",
                  "6.12", "6.13", "6.14", "6.15", "6.16", "6.17"]
# Tables with geographic COLUMNS (transpose case)
GEO_COL_TABLES = ["6.1a", "6.1b", "6.2", "6.9", "6.11"]


def extract_geo_row_table(table_id: str) -> pd.DataFrame:
    """For tables where rows are regions and cols are variables."""
    xlsx_path = SRC / f"Tabela {table_id}.xlsx"
    if not xlsx_path.exists():
        print(f"  ⚠ {xlsx_path.name} not found, skipping")
        return pd.DataFrame()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    records = []
    for year_str in YEARS:
        if year_str not in wb.sheetnames:
            continue
        ws = wb[year_str]
        # First pass: find max column count
        max_col = ws.max_column
        # Iterate all rows; capture those whose col[0] is a non-empty string
        for row_idx in range(1, ws.max_row + 1):
            label = ws.cell(row_idx, 1).value
            if not isinstance(label, str):
                continue
            label = label.strip()
            if not label or label.startswith(("Tabela", "Fonte:", "Nota:")):
                continue
            # Capture all columns
            row_values = {"year": int(year_str), "row_label": label, "row_index": row_idx}
            for c in range(2, max_col + 1):
                v = ws.cell(row_idx, c).value
                row_values[f"c{c:02d}"] = v
            records.append(row_values)
    return pd.DataFrame(records)


def extract_geo_col_table(table_id: str) -> pd.DataFrame:
    """For tables 6.1a, 6.1b, 6.2, 6.9, 6.11 — geography is in COLUMNS, not rows.

    Extract all cells as long format: (year, row_label, col_label, value).
    """
    xlsx_path = SRC / f"Tabela {table_id}.xlsx"
    if not xlsx_path.exists():
        print(f"  ⚠ {xlsx_path.name} not found, skipping")
        return pd.DataFrame()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    records = []
    for year_str in YEARS:
        if year_str not in wb.sheetnames:
            continue
        ws = wb[year_str]
        for row_idx in range(1, ws.max_row + 1):
            row_label = ws.cell(row_idx, 1).value
            if not isinstance(row_label, str):
                continue
            row_label = row_label.strip()
            if not row_label or row_label.startswith(("Tabela", "Fonte:", "Nota:")):
                continue
            for col_idx in range(2, ws.max_column + 1):
                v = ws.cell(row_idx, col_idx).value
                if v is None:
                    continue
                records.append({
                    "year": int(year_str),
                    "row_label": row_label,
                    "row_index": row_idx,
                    "col_index": col_idx,
                    "value": v if isinstance(v, (int, float)) else None,
                    "value_str": str(v),
                })
    return pd.DataFrame(records)


def load_and_push(df: pd.DataFrame, table_name: str) -> None:
    if df.empty:
        print(f"  ⚠ {table_name} empty, skipping write")
        return
    out_path = OUT / f"{table_name}.parquet"
    local_con = duckdb.connect()
    local_con.register("df_data", df)
    local_con.execute(
        f"COPY df_data TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)"
    )
    size_kb = out_path.stat().st_size / 1024
    print(f"  ✓ {out_path.name} — {len(df):,} rows, {size_kb:.1f} KB")

    token = os.environ.get("MOTHERDUCK_TOKEN")
    if token:
        con = duckdb.connect(f"md:atana?motherduck_token={token}")
        con.execute("CREATE SCHEMA IF NOT EXISTS atana.ibge_pnadc")
        con.register("df_data", df)
        con.execute(
            f"CREATE OR REPLACE TABLE atana.ibge_pnadc.{table_name} AS SELECT * FROM df_data"
        )
        n = con.execute(f"SELECT COUNT(*) FROM atana.ibge_pnadc.{table_name}").fetchone()[0]
        print(f"    Synced atana.ibge_pnadc.{table_name} ({n:,} rows)")


def main() -> None:
    if not SRC.exists():
        sys.exit(f"Source folder not found: {SRC}")
    print("Extracting IBGE PNADC tables...")
    print()
    print("Geographic-row tables (regions as rows):")
    for tid in GEO_ROW_TABLES:
        table_name = "tab_" + tid.replace(".", "_")
        df = extract_geo_row_table(tid)
        load_and_push(df, table_name)
    print()
    print("Geographic-column tables (regions as columns — long format):")
    for tid in GEO_COL_TABLES:
        table_name = "tab_" + tid.replace(".", "_")
        df = extract_geo_col_table(tid)
        load_and_push(df, table_name)


if __name__ == "__main__":
    main()
