"""PNAB (Política Nacional Aldir Blanc) → schema atana.pnab.

Phase PNAB ingest (2026-06-13). Four MinC transparency-portal xlsx exports
(DFD/SEFIC, atualização mensal) → three curated Parquet tables + a derived
governance table. Idempotent: inline-parsed → DuckDB COPY (ZSTD), stable sort,
byte-identical reruns. MotherDuck sync is a manual checkpoint (NEW schema).

REAL SCHEMAS (verified 2026-06-13 — diverge from the planning prompt):
  execucao_financeira : sheet "Execução Financeira"; header r4, DAX r5, data r6+
                        9 cols; NO cnpj — keys on cod_ibge / uf / ente(name).
  PAR Ciclo 1         : sheet "PAR - Ciclo 1"; header r3, DAX r4, data r5+
                        26 cols; conselho/plano/fundo = cols 23/24/25.
  PAR Ciclo 2         : sheet "Informações do PAR"; header r3, DAX r4, data r5+
                        16 cols, MULTI-SHEET workbook. Materially different
                        layout but carries the SAME analytical core (cod_ibge,
                        ente, cnpj, valor [col 12], conselho/plano/fundo
                        [cols 11/13/14]). Per prompt §13: harmonised into ONE
                        par_planos on the common columns + `ciclo` flag;
                        Ciclo-1-only rich text (responsável, meta-*, ações
                        afirmativas …) is NULL for Ciclo-2 rows. Documented in
                        docs/methodology/pnab_aldir_blanc.md §3.2.
  extratos            : sheet "ExtratoBancariofinal"; header r4, DAX r5, data r6+
                        8 cols (col2 blank); NO cod_ibge / NO cnpj — keys on
                        uf + ente(name) + recebedor.

CROSS-TABLE KEY: cod_ibge (execução ↔ par); extrato joins on uf+ente only.
Monetary values kept in BRL NOMINAL (not deflated — see methodology §4).
"""
import glob
import json
import re
from datetime import date
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "raw" / "pnab" / "_source"
OUT = REPO / "curated" / "pnab"
OUT.mkdir(parents=True, exist_ok=True)


def f(num):
    return float(num) if num is not None and num != "" else None


def s(x):
    return str(x).strip() if x is not None else None


def digits(x):
    return re.sub(r"\D", "", str(x)) if x is not None else None


def to_date(x):
    if x is None:
        return None
    if hasattr(x, "isoformat"):
        return x.date().isoformat() if hasattr(x, "date") else x.isoformat()
    t = str(x).strip()
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", t)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", t)
    return m.group(0) if m else None


def write(df_rows, table, sort_cols):
    df = df_rows if isinstance(df_rows, pd.DataFrame) else pd.DataFrame(df_rows)
    con = duckdb.connect()
    con.register("df_rows", df)
    con.execute("CREATE TABLE t AS SELECT * FROM df_rows")
    order = ", ".join(f'"{c}"' for c in sort_cols)
    con.execute(f"CREATE TABLE s AS SELECT * FROM t ORDER BY {order} NULLS LAST")
    p = OUT / f"{table}.parquet"
    con.execute(f"COPY s TO '{p}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    n = con.execute("SELECT count(*) FROM s").fetchone()[0]
    con.close()
    (OUT / f"{table}.meta.json").write_text(json.dumps({
        "table": f"pnab.{table}", "rows": n,
        "source": "MinC Portal de Dados da Cultura — dataset 'Implementação e "
                  "Execução da PNAB' (DFD/SEFIC); LC 195/2022 + Decreto 11.453/2023",
        "currency": "BRL nominal (not deflated — see methodology §4)",
        "etl_script": "etl/pnab__to_parquet.py", "etl_run_date": str(date.today()),
        "licence": "CC BY 4.0 (Portal de Dados da Cultura)",
    }, indent=2, ensure_ascii=False))
    print(f"  ✓ {table}: {n:,} rows")
    return n


def src(pat):
    return [x for x in glob.glob(str(SRC / "*.xlsx")) if pat in x][0]


# ── 1. execucao_financeira (Ciclo 1) ─────────────────────────────────────
def etl_execucao():
    wb = openpyxl.load_workbook(src("execucao-financeira"), read_only=True, data_only=True)
    ws = wb["Execução Financeira"]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 6:
            continue
        if r[0] is None:
            break
        rows.append({
            "tipo_ente": s(r[0]), "cod_ibge": s(r[1]), "uf": s(r[2]), "ente": s(r[3]),
            "total_recebido_brl": f(r[4]), "rendimentos_brl": f(r[5]),
            "saldo_brl": f(r[6]), "valor_gasto_brl": f(r[7]),
            "pct_gasto_decimal": f(r[8]), "ciclo": 1, "data_atualizacao": "2025-11-01",
        })
    wb.close()
    return write(rows, "execucao_financeira", ["cod_ibge"])


# ── 2. par_planos (Ciclo 1 ∪ Ciclo 2, harmonised) ────────────────────────
PAR_COLS = ["cod_ibge", "uf", "ente", "cnpj_ente", "cnpj_fundo", "nome_fundo",
            "codigo_plano_acao", "id_plano_acao", "valor_plano_brl", "data_envio_par",
            "situacao", "ano_par", "responsavel_1_nome", "responsavel_1_cargo",
            "responsavel_2_nome", "responsavel_2_cargo", "meta_acoes_gerais_text",
            "meta_acoes_gerais_outros_text", "meta_pncv_text",
            "meta_custo_operacional_5pct_text", "meta_custo_operacional_outros_text",
            "acoes_afirmativas_text", "atividades_perifericas_text",
            "participacao_social_text", "consulta_publica_text",
            "tem_conselho_cultura", "tem_plano_cultura", "tem_fundo_cultura", "ciclo"]


def blank_par():
    return {c: None for c in PAR_COLS}


def etl_par():
    rows = []
    # Ciclo 1 — 26 cols, header r3 / data r5+
    wb = openpyxl.load_workbook(src("paar-ciclo-1"), read_only=True, data_only=True)
    ws = wb["PAR - Ciclo 1"]
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        if r[0] is None:
            break
        d = blank_par()
        d.update(cod_ibge=s(r[0]), uf=s(r[1]), ente=s(r[2]), cnpj_ente=digits(r[3]),
                 cnpj_fundo=digits(r[4]), nome_fundo=s(r[5]), codigo_plano_acao=s(r[6]),
                 id_plano_acao=s(r[7]), valor_plano_brl=f(r[8]), data_envio_par=to_date(r[9]),
                 responsavel_1_nome=s(r[10]), responsavel_1_cargo=s(r[11]),
                 responsavel_2_nome=s(r[12]), responsavel_2_cargo=s(r[13]),
                 meta_acoes_gerais_text=s(r[14]), meta_acoes_gerais_outros_text=s(r[15]),
                 meta_pncv_text=s(r[16]), meta_custo_operacional_5pct_text=s(r[17]),
                 meta_custo_operacional_outros_text=s(r[18]), acoes_afirmativas_text=s(r[19]),
                 atividades_perifericas_text=s(r[20]), participacao_social_text=s(r[21]),
                 consulta_publica_text=s(r[22]), tem_conselho_cultura=s(r[23]),
                 tem_plano_cultura=s(r[24]), tem_fundo_cultura=s(r[25]), ciclo=1)
        rows.append(d)
    wb.close()
    n1 = len(rows)
    # Ciclo 2 — sheet "Informações do PAR", 16 cols, header r3 / data r5+
    wb = openpyxl.load_workbook(src("Ciclo 2"), read_only=True, data_only=True)
    ws = wb["Informações do PAR"]
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        if r[0] is None:
            break
        d = blank_par()
        d.update(cod_ibge=s(r[0]), uf=s(r[1]), ente=s(r[2]), cnpj_ente=digits(r[3]),
                 codigo_plano_acao=s(r[4]), ano_par=s(r[5]), data_envio_par=to_date(r[6]),
                 id_plano_acao=s(r[7]), situacao=s(r[8]), consulta_publica_text=s(r[9]),
                 valor_plano_brl=f(r[12]), tem_conselho_cultura=s(r[11]),
                 tem_plano_cultura=s(r[13]), tem_fundo_cultura=s(r[14]), ciclo=2)
        rows.append(d)
    wb.close()
    print(f"    (Ciclo 1: {n1:,} · Ciclo 2: {len(rows)-n1:,})")
    return write(rows, "par_planos", ["ciclo", "cod_ibge", "codigo_plano_acao"])


# ── 3. governanca_entes (derived from par_planos Ciclo 1) ────────────────
def etl_governanca():
    con = duckdb.connect()
    p = OUT / "par_planos.parquet"
    yes = "= 'Sim'"
    df = con.execute(f"""
        SELECT cod_ibge, uf, ente, cnpj_ente,
          CASE WHEN tem_conselho_cultura {yes} THEN TRUE
               WHEN tem_conselho_cultura = 'Não' THEN FALSE ELSE NULL END AS tem_conselho_cultura_bool,
          CASE WHEN tem_plano_cultura {yes} THEN TRUE
               WHEN tem_plano_cultura = 'Não' THEN FALSE ELSE NULL END AS tem_plano_cultura_bool,
          CASE WHEN tem_fundo_cultura {yes} THEN TRUE
               WHEN tem_fundo_cultura = 'Não' THEN FALSE ELSE NULL END AS tem_fundo_cultura_bool,
          (CASE WHEN tem_conselho_cultura {yes} THEN 1 ELSE 0 END)
          + (CASE WHEN tem_plano_cultura {yes} THEN 1 ELSE 0 END)
          + (CASE WHEN tem_fundo_cultura {yes} THEN 1 ELSE 0 END) AS escore_governanca,
          1 AS ciclo
        FROM read_parquet('{p}')
        WHERE ciclo = 1
        QUALIFY row_number() OVER (PARTITION BY cod_ibge ORDER BY codigo_plano_acao) = 1
        ORDER BY cod_ibge
    """).fetchdf()
    con.close()
    return write(df, "governanca_entes", ["cod_ibge"])


# ── 4. extratos_bancarios (Ciclo 1, streamed) ────────────────────────────
def etl_extratos():
    wb = openpyxl.load_workbook(src("extrato-bancario"), read_only=True, data_only=True)
    ws = wb["ExtratoBancariofinal"]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 6:
            continue
        if r[0] is None and r[1] is None:
            continue
        rows.append({
            "uf": s(r[0]), "ente": s(r[1]), "recebedor_recurso": s(r[3]),
            "descricao_lancamento": s(r[4]), "tipo_operacao": s(r[5]),
            "data_movimentacao": to_date(r[6]), "valor_brl": f(r[7]), "ciclo": 1,
        })
    wb.close()
    return write(rows, "extratos_bancarios",
                 ["uf", "ente", "data_movimentacao", "valor_brl"])


if __name__ == "__main__":
    print("PNAB ingest →", OUT)
    etl_execucao()
    etl_par()
    etl_governanca()
    etl_extratos()
    print("done — MotherDuck sync is manual (NEW schema atana.pnab).")
