"""Convert the Cuenta Satélite de Cultura de Costa Rica (CSCCR) — cultural
external-trade module — from the BCCR summary workbook to Parquet.

Phase 3d of the Atana Data LATAM expansion. Fourth national source, after
Mexico (`atana.inegi`), Colombia (`atana.dane`) and Argentina (`atana.sinca`).

WHAT THIS INGESTS
-----------------
The CSCCR is built by the CICSC consortium (Ministerio de Cultura y Juventud,
Banco Central de Costa Rica, INEC, Programa Estado de la Nación, CONARE). Unlike
Mexico and Colombia — where cultural trade had to be dug out of the supply-use
tables — the CSCCR publishes a dedicated, consolidated trade sheet.

Source file (downloaded manually from the BCCR CSC portal):
    raw/cr_bccr/_source/Resumen de indicadores.xlsx
        → sheet "Comercio exterior cultural" — exports & imports of 4 cultural
          sectors (Editorial, Publicidad, Audiovisual, Música), 2010–2024,
          millions of colones (CRC)
        → sheet "TC" — annual colón/USD exchange rate, 2010–2021

⚠️ COVERAGE BREAK. The trade sheet covers all four sectors only for 2010–2021.
From 2022 onward only the Editorial sector is reported (the other three show
`n.d.`), and the published year totals collapse to Editorial-only. The
`full_sector_coverage` column flags this: TRUE for 2010–2021, FALSE for
2022–2024. Totals for 2022–2024 are NOT comparable with 2010–2021 totals.

Output (long format):
    raw/cr_bccr/csc_comercio.parquet            (+ .meta.json)
    raw/cr_bccr/_reference/fx_crc_usd_annual.parquet  (+ .meta.json)

Idempotent. openpyxl read → list-of-dicts → DuckDB COPY to Parquet (no pyarrow).
MotherDuck push gated behind a token (env var or .motherduck_token file).

Usage:
    python etl/cr_bccr__csc_to_parquet.py
"""
import hashlib
import json
import os
from datetime import date
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "raw" / "cr_bccr"
REF = OUT / "_reference"
SRC = OUT / "_source"
OUT.mkdir(parents=True, exist_ok=True)
REF.mkdir(parents=True, exist_ok=True)

SOURCE_FILE = SRC / "Resumen de indicadores.xlsx"
TRADE_SHEET = "Comercio exterior cultural"
TC_SHEET = "TC"

YEAR_HEADER_ROW = 10           # row 10: "DESCRIPCIÓN | 2010 | 2011 | ... | 2024"
YEAR_FIRST_COL = 2             # 2010 in column B
TRADE_FIRST_ROW = 11           # "Exportaciones" total
TRADE_LAST_ROW = 20            # last sector row

SECTOR_LABELS = {"editorial", "publicidad", "audiovisual", "música", "musica"}
FULL_COVERAGE_THROUGH = 2021   # 2022+ is Editorial-only — see header note

# colón/USD for 2022–2024 — the CSCCR 'TC' sheet stops at 2021.
# Annual averages: 2022 & 2024 from CEIC / World Bank PA.NUS.FCRF; 2023
# transcribed approximate (the colón appreciated sharply that year).
# ⚠️ Re-verify against the live BCCR / World Bank series before publication use.
FX_SUPPLEMENT = {2022: 646.899, 2023: 542.5, 2024: 515.561}
FX_SUPPLEMENT_SOURCE = ("CEIC / World Bank PA.NUS.FCRF annual average — "
                        "transcribed 2026-05-22 (2023 approximate); re-verify")
FX_SHEET_SOURCE = "CSCCR 'TC' sheet — Banco Central de Costa Rica, annual average"


def _num(value):
    """Coerce a cell to float; 'n.d.' / blanks / non-numeric → None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    v = str(value).strip().lower()
    if v in ("", "n.d.", "nd", "n/d", "-", "..."):
        return None
    try:
        return float(v.replace(",", ""))
    except ValueError:
        return None


def extract_fx(wb) -> dict:
    """Read the 'TC' sheet (2010–2021) and supplement 2022–2024."""
    ws = wb[TC_SHEET]
    fx = {}
    for r in range(1, ws.max_row + 1):
        yr = ws.cell(r, 1).value
        rate = _num(ws.cell(r, 2).value)
        if isinstance(yr, (int, float)) and 2000 <= int(yr) <= 2035 and rate:
            fx[int(yr)] = {"rate": rate, "source": FX_SHEET_SOURCE}
    for yr, rate in FX_SUPPLEMENT.items():
        fx.setdefault(yr, {"rate": rate, "source": FX_SUPPLEMENT_SOURCE})
    return fx


def extract_comercio(wb, fx: dict) -> pd.DataFrame:
    """Parse the 'Comercio exterior cultural' sheet → long-format trade rows."""
    ws = wb[TRADE_SHEET]
    # year columns
    years = []
    for c in range(YEAR_FIRST_COL, ws.max_column + 1):
        v = ws.cell(YEAR_HEADER_ROW, c).value
        if isinstance(v, (int, float)) and 2000 <= int(v) <= 2035:
            years.append((int(v), c))

    rows = []
    current_flow = None
    for r in range(TRADE_FIRST_ROW, TRADE_LAST_ROW + 1):
        label = ws.cell(r, 1).value
        if not isinstance(label, str) or not label.strip():
            continue
        name = label.strip()
        low = name.lower()
        if low.startswith("exportacion"):
            current_flow, sector = "exportacion", "Total"
        elif low.startswith("importacion"):
            current_flow, sector = "importacion", "Total"
        elif low in SECTOR_LABELS:
            sector = name
        else:
            continue
        for year, col in years:
            crc = _num(ws.cell(r, col).value)
            rate = fx.get(year, {}).get("rate")
            usd = round(crc / rate, 6) if (crc is not None and rate) else None
            rows.append(dict(
                year=year,
                sector=sector,
                flow=current_flow,
                value_crc_million=crc,
                value_usd_million=usd,
                fx_rate_crc_per_usd=rate,
                full_sector_coverage=(year <= FULL_COVERAGE_THROUGH),
            ))
    return pd.DataFrame(rows).sort_values(
        ["flow", "year", "sector"]).reset_index(drop=True)


def write_parquet(df: pd.DataFrame, table_name: str, out_dir: Path) -> Path:
    out_path = out_dir / f"{table_name}.parquet"
    con = duckdb.connect()
    con.register("df_data", df)
    con.execute(f"COPY df_data TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    size_kb = out_path.stat().st_size / 1024
    print(f"  ✓ {out_path.relative_to(REPO_ROOT)} — {len(df):,} rows, {size_kb:.1f} KB")
    return out_path


def write_meta(out_path: Path, source_files: list, description: str,
               extra: dict | None = None) -> None:
    srcs = []
    for sf in source_files:
        p = Path(sf)
        h = hashlib.sha256(p.read_bytes()).hexdigest() if p.exists() else None
        srcs.append({"file": p.name, "sha256": h})
    meta = {
        "table": out_path.stem,
        "description": description,
        "source": "Cuenta Satélite de Cultura de Costa Rica (CSCCR) — CICSC "
                   "(MCJ + BCCR + INEC + PEN + CONARE); data hosted by the "
                   "Banco Central de Costa Rica",
        "source_url": "https://www.bccr.fi.cr/indicadores-economicos/"
                      "cuentas-tematicas/cuenta-satelite-de-cultura",
        "source_files": srcs,
        "fetch_date": "2026-05-22",
        "etl_script": "etl/cr_bccr__csc_to_parquet.py",
        "etl_run_date": str(date.today()),
        "licence": "open data — Costa Rican official statistics (BCCR / MCJ)",
    }
    if extra:
        meta.update(extra)
    out_path.with_suffix(".meta.json").write_text(
        json.dumps(meta, indent=2, ensure_ascii=False))
    print(f"  ✓ {out_path.with_suffix('.meta.json').relative_to(REPO_ROOT)}")


def maybe_push(df: pd.DataFrame, schema: str, table: str) -> None:
    """Push to MotherDuck if a valid token is available (MOTHERDUCK_TOKEN env
    var or a gitignored .motherduck_token file; must be a JWT). Off by default."""
    def _jwt(t) -> str:
        t = (t or "").strip()
        return t if (t.startswith("eyJ") and t.count(".") == 2) else ""
    token = _jwt(os.environ.get("MOTHERDUCK_TOKEN"))
    if not token:
        tf = REPO_ROOT / ".motherduck_token"
        token = _jwt(tf.read_text()) if tf.exists() else ""
    if not token:
        print(f"  · MotherDuck push skipped for atana.{schema}.{table} — no valid "
              f"token. Put a real MotherDuck token (a JWT: starts 'eyJ', two dots) "
              f"in {REPO_ROOT}/.motherduck_token, then re-run.")
        return
    con = duckdb.connect(f"md:atana?motherduck_token={token}")
    con.execute(f"CREATE SCHEMA IF NOT EXISTS atana.{schema}")
    con.register("df_data", df)
    con.execute(
        f"CREATE OR REPLACE TABLE atana.{schema}.{table} AS SELECT * FROM df_data")
    n = con.execute(f"SELECT COUNT(*) FROM atana.{schema}.{table}").fetchone()[0]
    print(f"  ✓ Synced atana.{schema}.{table} ({n:,} rows)")


def main() -> None:
    if not SOURCE_FILE.exists():
        raise SystemExit(f"Source file not found: {SOURCE_FILE}")
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

    print("Building FX reference table...")
    fx = extract_fx(wb)
    fx_df = pd.DataFrame(
        [{"year": y, "fx_rate_crc_per_usd": d["rate"], "source": d["source"]}
         for y, d in sorted(fx.items())])
    fx_path = write_parquet(fx_df, "fx_crc_usd_annual", REF)
    write_meta(fx_path, [SOURCE_FILE], "Annual-average CRC/USD exchange rate, "
               "used to derive the USD column of cr_bccr.csc_comercio. 2010–2021 "
               "from the CSCCR 'TC' sheet (BCCR); 2022–2024 supplemented from "
               "World Bank / CEIC.",
               extra={"source": f"{FX_SHEET_SOURCE}; {FX_SUPPLEMENT_SOURCE}"})

    print("Extracting CSCCR cultural-trade rows...")
    com = extract_comercio(wb, fx)
    com_path = write_parquet(com, "csc_comercio", OUT)
    write_meta(com_path, [SOURCE_FILE],
               "Costa Rica cultural foreign trade — exports and imports of 4 "
               "cultural sectors (Editorial, Publicidad, Audiovisual, Música), "
               "2010–2024, in millions of colones (CRC). USD column ETL-derived.",
               extra={"grain": "year × sector × flow",
                      "currency": "CRC million (value_crc_million); USD million "
                                  "is derived, see fx_crc_usd_annual",
                      "coverage_break": "Full 4-sector coverage 2010–2021 only; "
                      "2022–2024 is Editorial-only (other sectors n.d.) and the "
                      "year totals collapse to Editorial. See full_sector_coverage."})

    maybe_push(fx_df, "cr_bccr", "fx_crc_usd_annual")
    maybe_push(com, "cr_bccr", "csc_comercio")
    print("Done.")


if __name__ == "__main__":
    main()
