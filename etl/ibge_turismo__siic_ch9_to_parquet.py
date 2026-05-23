"""SIIC Chapter 9 — Turismo de lazer, cultura e natureza → Parquet.

Phase 4b of the Atana Data expansion. Partially closes the FCS transversal
domain 'Social participation': the PNAD Contínua leisure-tourism supplement
records leisure travel by type — including a 'Cultura e gastronomia' category —
which a foreign-trade module cannot see.

⚠️ PROXY, NOT FULL MEASUREMENT. Chapter 9 is a leisure-travel proxy for the FCS
'Social participation' domain. Brazil has no continuous national survey of
cultural practice (museum / theatre / cinema attendance, reading). The crosswalk
rows for this domain carry mapping_confidence = 'approximate'. See the
methodology note and phase4_scoping.md §A.4.

WHAT THIS INGESTS
-----------------
Source (already in the workspace — SIIC 'Informações Culturais', 2024 edition):
    9_turismo_de_lazer_cultura_e_natureza/
        Tabela 9.1.xlsx .. Tabela 9.5.xlsx — one xlsx per SIIC table, each with
        a sheet per reference year (2021, 2023, 2024) and a paired '<year> (CV)'
        sheet of IBGE CV reliability codes.

The chapter's tables carry a multi-row IBGE header; this ETL preserves the
sheet's column structure faithfully (the `ibge_pnadc` precedent): one row per
labelled spreadsheet row, original columns kept as c02, c03, …. Column meaning
is documented in docs/methodology/ibge_turismo_siic_ch9.md.

OUTPUT, one Parquet per source table (all year + CV sheets stacked):
    raw/ibge_turismo/tab_9_1.parquet .. tab_9_5.parquet (+ .meta.json)
    grain: table_id × year × is_cv × row_label

Idempotent. openpyxl read → list-of-dicts → DuckDB COPY to Parquet (no pyarrow).
MotherDuck push gated behind a token (skipped when ATANA_ETL_SKIP_PUSH is set).
Schema: atana.ibge_turismo.

Usage:
    python etl/ibge_turismo__siic_ch9_to_parquet.py
"""
import glob
import hashlib
import json
import os
import re
from datetime import date
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "raw" / "ibge_turismo"
OUT.mkdir(parents=True, exist_ok=True)

_WS_CANDIDATES = [
    "/Users/joaoroque/Documents/Cultural production - book/"
    "Dados da Economia Cultural no Brasil",
    *sorted(glob.glob("/sessions/*/mnt/Dados da Economia Cultural no Brasil")),
]
WORKSPACE = next((Path(p) for p in _WS_CANDIDATES if Path(p).exists()), None)
SRC_DIR = (WORKSPACE / "9_turismo_de_lazer_cultura_e_natureza"
           if WORKSPACE else None)

TABLE_IDS = ["9.1", "9.2", "9.3", "9.4", "9.5"]
SHEET_RE = re.compile(r"^(\d{4})\s*(\(CV\))?$")
SKIP_PREFIXES = ("Tabela", "Fonte", "Nota")


def extract_table(path: Path, table_id: str) -> pd.DataFrame:
    """One xlsx (all year + CV sheets) → faithful wide rows."""
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        m = SHEET_RE.match(name.strip())
        if m:
            sheets.append((int(m.group(1)), bool(m.group(2)), wb[name]))

    max_col = 2
    for _, _, ws in sheets:
        for r in range(1, ws.max_row + 1):
            for c in range(ws.max_column, max_col, -1):
                if ws.cell(r, c).value is not None:
                    max_col = c
                    break

    records = []
    for year, is_cv, ws in sheets:
        for r in range(1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if not isinstance(label, str) or not label.strip():
                continue
            label = label.strip()
            if label.startswith(SKIP_PREFIXES):
                continue
            rec = {"table_id": table_id, "year": year, "is_cv": is_cv,
                   "row_index": r, "row_label": label}
            for c in range(2, max_col + 1):
                rec[f"c{c:02d}"] = ws.cell(r, c).value
            records.append(rec)
    wb.close()
    return pd.DataFrame(records).sort_values(
        ["is_cv", "year", "row_index"]).reset_index(drop=True)


def write_parquet(df: pd.DataFrame, table_name: str) -> Path:
    out_path = OUT / f"{table_name}.parquet"
    con = duckdb.connect()
    con.register("df_data", df)
    con.execute(f"COPY df_data TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    size_kb = out_path.stat().st_size / 1024
    print(f"  ✓ {out_path.relative_to(REPO_ROOT)} — {len(df):,} rows, {size_kb:.1f} KB")
    return out_path


def write_meta(out_path: Path, table_id: str, src_file: Path,
               years: list) -> None:
    h = hashlib.sha256(src_file.read_bytes()).hexdigest() if src_file.exists() else None
    meta = {
        "table": out_path.stem,
        "description": f"SIIC chapter 9, table {table_id} — PNAD Contínua "
                       "leisure-tourism supplement (leisure travel by type, "
                       "incl. 'Cultura e gastronomia'), cultural-participation "
                       "proxy. Year + CV sheets stacked; IBGE column structure "
                       "preserved as c02, c03, ….",
        "source": "IBGE — Sistema de Informações e Indicadores Culturais "
                   "(SIIC), 'Informações Culturais' 2024 edition, chapter 9 "
                   "'Turismo de lazer, cultura e natureza'",
        "source_url": "https://www.ibge.gov.br/estatisticas/sociais/cultura.html",
        "source_files": [{"file": src_file.name, "sha256": h}],
        "fetch_date": "2026-05-23",
        "etl_script": "etl/ibge_turismo__siic_ch9_to_parquet.py",
        "etl_run_date": str(date.today()),
        "licence": "open data — IBGE official statistics",
        "grain": "table_id × year × is_cv × row_label",
        "years": years,
        "notes": "is_cv = TRUE rows hold IBGE CV reliability codes (A best .. "
                 "E suppress) mirroring the data sheet. Multi-row IBGE headers "
                 "are not captured as rows — see the methodology note for the "
                 "c02… column meanings. PROXY domain: leisure-travel data, an "
                 "approximate proxy for FCS 'Social participation'.",
    }
    out_path.with_suffix(".meta.json").write_text(
        json.dumps(meta, indent=2, ensure_ascii=False))
    print(f"  ✓ {out_path.with_suffix('.meta.json').relative_to(REPO_ROOT)}")


def maybe_push(df: pd.DataFrame, schema: str, table: str) -> None:
    """Push to MotherDuck if a valid JWT token is available.

    Skipped entirely when ATANA_ETL_SKIP_PUSH is set — used for build-only /
    sandbox runs where the Parquet is produced but the sync stays with João.
    """
    if os.environ.get("ATANA_ETL_SKIP_PUSH"):
        print(f"  · push skipped for atana.{schema}.{table} (ATANA_ETL_SKIP_PUSH)")
        return

    def _jwt(t) -> str:
        t = (t or "").strip()
        return t if (t.startswith("eyJ") and t.count(".") == 2) else ""
    token = _jwt(os.environ.get("MOTHERDUCK_TOKEN"))
    if not token:
        tf = REPO_ROOT / ".motherduck_token"
        token = _jwt(tf.read_text()) if tf.exists() else ""
    if not token:
        print(f"  · MotherDuck push skipped for atana.{schema}.{table} — no "
              f"valid token. Put a JWT in {REPO_ROOT}/.motherduck_token, re-run.")
        return
    con = duckdb.connect(f"md:atana?motherduck_token={token}")
    con.execute(f"CREATE SCHEMA IF NOT EXISTS atana.{schema}")
    con.register("df_data", df)
    con.execute(
        f"CREATE OR REPLACE TABLE atana.{schema}.{table} AS SELECT * FROM df_data")
    n = con.execute(f"SELECT COUNT(*) FROM atana.{schema}.{table}").fetchone()[0]
    print(f"  ✓ Synced atana.{schema}.{table} ({n:,} rows)")


def main() -> None:
    if not SRC_DIR or not SRC_DIR.exists():
        raise SystemExit(f"Source folder not found (workspace: {WORKSPACE})")
    print(f"Reading SIIC chapter 9 from {SRC_DIR.name}")
    for table_id in TABLE_IDS:
        src_file = SRC_DIR / f"Tabela {table_id}.xlsx"
        if not src_file.exists():
            print(f"  ⚠ {src_file.name} missing — skipped")
            continue
        df = extract_table(src_file, table_id)
        table_name = "tab_" + table_id.replace(".", "_")
        years = sorted(int(y) for y in df["year"].unique())
        out_path = write_parquet(df, table_name)
        write_meta(out_path, table_id, src_file, years)
        maybe_push(df, "ibge_turismo", table_name)
    print("Done.")


if __name__ == "__main__":
    main()
