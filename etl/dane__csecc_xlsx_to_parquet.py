"""Convert DANE Cuenta Satélite de Economía Cultural y Creativa (CSECC) —
cultural external-trade module — from the CSECC balance xlsx to Parquet.

Phase 3b of the Atana Data LATAM expansion (Phase 3 — multi-country).
Second national source after Mexico's INEGI CSCM (`atana.inegi`).

WHAT THIS INGESTS
-----------------
DANE publishes, alongside each CSECC release, a workbook of product-level
"Balances oferta utilización" (supply-use balances), one cuadro per cultural
product. Cultural imports/exports appear as rows *inside* those balances —
`Importaciones` on the supply side, `Exportaciones a precio comprador` on the
use side. Like Mexico's CSCM, the CSECC has no standalone balance-of-payments
module and no bilateral partner detail; trade lives inside the supply-use
tables.

Source file (downloaded manually from dane.gov.co, CSECC 2022-2024pr release):
    raw/dane/_source/anex-CSECC-balanceOferUtilizacion-2024pr.xlsx
        — 35 cuadros (one product each), 2014–2024pr, current prices, COP million
        — grouped into DANE's 3 areas:
            cuadros 1–13  → Artes y patrimonio
            cuadros 14–29 → Industrias culturales
            cuadros 30–35 → Creaciones funcionales

Output (long format):
    raw/dane/csecc_comercio.parquet                  (+ .meta.json sidecar)
    raw/dane/_reference/fx_cop_usd_annual.parquet    (+ .meta.json sidecar)

Idempotent. Mirrors etl/inegi__csc_xlsx_to_parquet.py — openpyxl read →
list-of-dicts → pandas DataFrame → DuckDB COPY to Parquet (no pyarrow). The
MotherDuck push is gated behind MOTHERDUCK_TOKEN and is NOT run unless set.

Usage:
    python etl/dane__csecc_xlsx_to_parquet.py
    # optional, only when explicitly authorised:
    export MOTHERDUCK_TOKEN="..." && python etl/dane__csecc_xlsx_to_parquet.py
"""
import hashlib
import json
import os
import sys
from datetime import date
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "raw" / "dane"
REF = OUT / "_reference"
SRC = OUT / "_source"
OUT.mkdir(parents=True, exist_ok=True)
REF.mkdir(parents=True, exist_ok=True)

SOURCE_FILE = SRC / "anex-CSECC-balanceOferUtilizacion-2024pr.xlsx"

N_CUADROS = 35
YEAR_HEADER_ROW = 13          # row 13 holds the year labels
YEAR_FIRST_COL = 2            # column B
PRODUCT_NAME_ROW = 6          # row 6 holds the product name

# Cuadro number → DANE area (verified against the workbook's Índice sheet).
def _area_for_cuadro(n: int) -> str:
    if 1 <= n <= 13:
        return "Artes y patrimonio"
    if 14 <= n <= 29:
        return "Industrias culturales"
    if 30 <= n <= 35:
        return "Creaciones funcionales"
    return "Desconocida"

# ---------------------------------------------------------------------------
# FX reference — annual-average COP per USD.
# Source: World Bank Open Data, indicator PA.NUS.FCRF ("Official exchange rate,
# LCU per US$, period average"; underlying source IMF International Financial
# Statistics). Values transcribed 2026-05-22.
# ⚠️ Documented convenience series for the *derived* USD column — the CSECC
# itself publishes no USD figures. Re-verify against the live World Bank /
# Banco de la República series before any publication relies on the USD numbers.
# ---------------------------------------------------------------------------
FX_COP_PER_USD = {
    2014: 2000.68, 2015: 2741.88, 2016: 3053.42, 2017: 2951.33,
    2018: 2956.44, 2019: 3281.09, 2020: 3693.27, 2021: 3743.59,
    2022: 4256.18, 2023: 4325.92, 2024: 4071.51,
}
FX_SOURCE = ("World Bank Open Data PA.NUS.FCRF (Official exchange rate, "
             "LCU per US$, period average; IMF IFS) — transcribed 2026-05-22, "
             "re-verify before publication use")


def _num(value):
    """Coerce a cell to float; blanks / non-numeric → None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    v = str(value).strip().replace(",", "")
    if v == "":
        return None
    try:
        return float(v)
    except ValueError:
        return None


def _parse_year(label):
    """'2014' → (2014,'definitivo'); '2023p' → (2023,'provisional');
    '2024pr' → (2024,'preliminar'). Returns None if not a year."""
    if label is None:
        return None
    s = str(label).strip()
    status = "definitivo"
    if s.endswith("pr"):
        status, s = "preliminar", s[:-2]
    elif s.endswith("p"):
        status, s = "provisional", s[:-1]
    if s.isdigit() and 2000 <= int(s) <= 2035:
        return int(s), status
    return None


def extract() -> list:
    """Parse the 35 product cuadros → long-format trade rows."""
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    rows = []
    for n in range(1, N_CUADROS + 1):
        sheet = f"Cuadro {n}"
        if sheet not in wb.sheetnames:
            print(f"  ! {sheet} missing — skipped")
            continue
        ws = wb[sheet]
        producto = str(ws.cell(PRODUCT_NAME_ROW, 1).value or "").strip()
        area = _area_for_cuadro(n)

        # year columns
        years = []
        for c in range(YEAR_FIRST_COL, ws.max_column + 1):
            parsed = _parse_year(ws.cell(YEAR_HEADER_ROW, c).value)
            if parsed:
                years.append({"year": parsed[0], "status": parsed[1], "col": c})

        # find the import / export rows by label (positions vary per cuadro)
        for r in range(YEAR_HEADER_ROW + 1, ws.max_row + 1):
            lab = ws.cell(r, 1).value
            if not isinstance(lab, str):
                continue
            low = lab.strip().lower()
            if low.startswith("importacion"):
                flow = "importacion"
            elif low.startswith("exportacion"):
                flow = "exportacion"
            else:
                continue
            for y in years:
                fx = FX_COP_PER_USD.get(y["year"])
                val = _num(ws.cell(r, y["col"]).value)
                usd = round(val / fx, 6) if (val is not None and fx) else None
                rows.append(dict(
                    year=y["year"],
                    area=area,
                    cuadro_num=n,
                    producto=producto,
                    flow=flow,
                    value_cop_million=val,
                    value_usd_million=usd,
                    fx_rate_cop_per_usd=fx,
                    year_status=y["status"],
                    source_concept=lab.strip(),
                ))
    return rows


def write_parquet(df: pd.DataFrame, table_name: str, out_dir: Path) -> Path:
    out_path = out_dir / f"{table_name}.parquet"
    con = duckdb.connect()
    con.register("df_data", df)
    con.execute(f"COPY df_data TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    size_kb = out_path.stat().st_size / 1024
    print(f"  ✓ {out_path.relative_to(REPO_ROOT)} — {len(df):,} rows, {size_kb:.1f} KB")
    return out_path


def write_meta(out_path: Path, source_files: list, description: str,
               extra: dict | None = None) -> None:
    srcs = []
    for sf in source_files:
        p = Path(sf)
        h = hashlib.sha256(p.read_bytes()).hexdigest() if p.exists() else None
        srcs.append({"file": p.name, "sha256": h})
    meta = {
        "table": out_path.stem,
        "description": description,
        "source": "DANE — Cuenta Satélite de Economía Cultural y Creativa "
                   "(CSECC), release 2022–2024pr",
        "source_url": "https://www.dane.gov.co/index.php/estadisticas-por-tema/"
                      "cuentas-nacionales/cuentas-satelite/"
                      "cuenta-satelite-de-cultura-en-colombia",
        "source_files": srcs,
        "fetch_date": "2026-05-22",
        "etl_script": "etl/dane__csecc_xlsx_to_parquet.py",
        "etl_run_date": str(date.today()),
        "licence": "DANE — open data (Colombian official statistics)",
    }
    if extra:
        meta.update(extra)
    out_path.with_suffix(".meta.json").write_text(
        json.dumps(meta, indent=2, ensure_ascii=False))
    print(f"  ✓ {out_path.with_suffix('.meta.json').relative_to(REPO_ROOT)}")


def maybe_push(df: pd.DataFrame, schema: str, table: str) -> None:
    """Push to MotherDuck if a valid token is available. The token is read from
    the MOTHERDUCK_TOKEN env var, or from a gitignored `.motherduck_token` file
    in the repo root. A MotherDuck token is a JWT — it starts with 'eyJ' and has
    exactly two dots; anything else (placeholders, stray command text) is
    rejected so it fails fast with a clear message. Off by default."""
    def _jwt(t) -> str:
        t = (t or "").strip()
        return t if (t.startswith("eyJ") and t.count(".") == 2) else ""
    token = _jwt(os.environ.get("MOTHERDUCK_TOKEN"))
    if not token:
        tf = REPO_ROOT / ".motherduck_token"
        token = _jwt(tf.read_text()) if tf.exists() else ""
    if not token:
        print(f"  · MotherDuck push skipped for atana.{schema}.{table} — no valid "
              f"token. Put a real MotherDuck token (a JWT: starts 'eyJ', two dots) "
              f"in {REPO_ROOT}/.motherduck_token, then re-run.")
        return
    con = duckdb.connect(f"md:atana?motherduck_token={token}")
    con.execute(f"CREATE SCHEMA IF NOT EXISTS atana.{schema}")
    con.register("df_data", df)
    con.execute(
        f"CREATE OR REPLACE TABLE atana.{schema}.{table} AS SELECT * FROM df_data")
    n = con.execute(f"SELECT COUNT(*) FROM atana.{schema}.{table}").fetchone()[0]
    print(f"  ✓ Synced atana.{schema}.{table} ({n:,} rows)")


def main() -> None:
    if not SOURCE_FILE.exists():
        sys.exit(f"Source file not found: {SOURCE_FILE}")

    # ── 1. FX reference table ────────────────────────────────────────────────
    print("Building FX reference table...")
    fx_df = pd.DataFrame(
        [{"year": y, "fx_rate_cop_per_usd": r, "source": FX_SOURCE}
         for y, r in sorted(FX_COP_PER_USD.items())])
    fx_path = write_parquet(fx_df, "fx_cop_usd_annual", REF)
    write_meta(fx_path, [], "Annual-average COP/USD exchange rate, used to "
               "derive the USD column of dane.csecc_comercio. NOT a DANE "
               "figure — a documented convenience series.",
               extra={"source": FX_SOURCE,
                      "source_url": "https://datos.bancomundial.org/indicador/PA.NUS.FCRF?locations=CO"})

    # ── 2. csecc_comercio — the trade module ─────────────────────────────────
    print("Extracting CSECC cultural-trade rows...")
    rows = extract()
    df = pd.DataFrame(rows).sort_values(
        ["year", "area", "cuadro_num", "flow"]).reset_index(drop=True)
    n_cuadros_with_trade = df["cuadro_num"].nunique()
    print(f"  · {n_cuadros_with_trade} of {N_CUADROS} cuadros carry trade rows")
    com_path = write_parquet(df, "csecc_comercio", OUT)
    write_meta(com_path, [SOURCE_FILE],
               "Cultural imports (Importaciones) and exports (Exportaciones) "
               "from the DANE CSECC product-level supply-use balances, by "
               "product × area × year, 2014–2024, in million COP (current "
               "prices). USD column is ETL-derived.",
               extra={"grain": "year × cuadro(product) × flow",
                      "currency": "COP million (value_cop_million); USD million "
                                  "is derived, see fx_cop_usd_annual",
                      "note_no_bop_module": "The CSECC has no balance-of-payments "
                      "module; these figures are the import/export rows of the "
                      "product-level supply-use balances. No bilateral partner "
                      "detail. Imports valued CIF / precios básicos, exports a "
                      "precio comprador (see source_concept column)."})

    # ── 3. MotherDuck (gated, off by default) ────────────────────────────────
    maybe_push(fx_df, "dane", "fx_cop_usd_annual")
    maybe_push(df, "dane", "csecc_comercio")
    print("Done.")


if __name__ == "__main__":
    main()
