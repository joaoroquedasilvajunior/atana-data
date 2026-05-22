"""Convert INEGI Cuenta Satélite de la Cultura de México (CSCM) — cultural
external-trade module — from the CSCM tabulado xlsx files to Parquet.

Phase 3a.1 of the Atana Data LATAM expansion (Phase 3 — multi-country).

RE-SCOPE NOTE
-------------
The Phase 3a scoping memo (2026-05-16) assumed the CSCM publishes a dedicated
balance-of-payments module (`Cuadro 3/4/5 — Comercio exterior cultural`, pesos
+ USD). Web verification (2026-05-22) established that no such module exists:
the CSCM is peso-only and has no BoP cuadro. Cultural imports/exports live
*inside* the Cuadros de Oferta y Utilización (supply-use tables) — Importaciones
C.I.F. on the supply side, Exportaciones on the use side.

This script therefore ingests the import/export columns of:
    CSCM_31.xlsx  — Valores corrientes, áreas generales y específicas, O&U
    CSCM_75.xlsx  — Valores constantes (precios de 2018), áreas gen./esp., O&U

Source files (downloaded manually from inegi.org.mx/programas/cultura/2018/
"Tabulados", and unzipped):
    raw/inegi/_source/tabulados_CSCM/CSCM_31.xlsx
    raw/inegi/_source/tabulados_CSCM/CSCM_75.xlsx

Output (long format, one Parquet per table):
    raw/inegi/csc_comercio.parquet                      (+ .meta.json sidecar)
    raw/inegi/_reference/fx_mxn_usd_annual.parquet      (+ .meta.json sidecar)

Idempotent: rerunning produces identical output. Follows the
etl/ibge_comex__xlsx_to_parquet.py pattern — openpyxl read → list-of-dicts →
pandas DataFrame → DuckDB COPY to Parquet (no pyarrow dependency). The
MotherDuck push is gated behind the MOTHERDUCK_TOKEN env var and is NOT run
unless that variable is set.

Usage:
    python etl/inegi__csc_xlsx_to_parquet.py
    # optional, only when explicitly authorised:
    export MOTHERDUCK_TOKEN="..." && python etl/inegi__csc_xlsx_to_parquet.py
"""
import hashlib
import json
import os
import sys
from datetime import date
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "raw" / "inegi"
REF = OUT / "_reference"
SRC = OUT / "_source" / "tabulados_CSCM"
OUT.mkdir(parents=True, exist_ok=True)
REF.mkdir(parents=True, exist_ok=True)

# The two CSCM oferta-utilización cuadros that carry the trade columns.
CUADROS = {
    "CSCM_31": {"file": "CSCM_31.xlsx", "price_basis": "corriente"},
    "CSCM_75": {"file": "CSCM_75.xlsx", "price_basis": "constante_2018"},
}

# Column offsets WITHIN each 10-column year block of the O&U cuadro.
# block: [Producción bruta, Importaciones C.I.F., Márgenes, O&U total,
#         Demanda intermedia, Consumo privado, Consumo gobierno,
#         Formación bruta de capital, Variación de existencias, Exportaciones]
OFFSET_IMPORTACION = 1   # Importaciones C.I.F.   (supply side)
OFFSET_EXPORTACION = 9   # Exportaciones          (use side)

# Indent (leading spaces in column A) → hierarchy level in the CSCM cuadros.
INDENT_TOTAL = 10        # "Total cultura"
INDENT_AREA_GENERAL = 15 # the 10 áreas generales
INDENT_AREA_ESPECIFICA = 20  # the áreas específicas

# ---------------------------------------------------------------------------
# FX reference — annual-average MXN per USD.
# Source: World Bank Open Data, indicator PA.NUS.FCRF ("Official exchange rate,
# LCU per US$, period average"; underlying source IMF International Financial
# Statistics). Values transcribed 2026-05-22.
# ⚠️ These are a documented convenience series for the *derived* USD column —
# the CSCM itself publishes no USD figures. Re-verify against the live World
# Bank / Banxico series before any publication relies on the USD numbers.
# ---------------------------------------------------------------------------
FX_MXN_PER_USD = {
    2008: 11.130, 2009: 13.514, 2010: 12.636, 2011: 12.423, 2012: 13.169,
    2013: 12.766, 2014: 13.292, 2015: 15.848, 2016: 18.664, 2017: 18.927,
    2018: 19.244, 2019: 19.263, 2020: 21.486, 2021: 20.272, 2022: 20.127,
    2023: 17.733, 2024: 18.330,
}
FX_SOURCE = ("World Bank Open Data PA.NUS.FCRF (Official exchange rate, "
             "LCU per US$, period average; IMF IFS) — transcribed 2026-05-22, "
             "re-verify before publication use")


def _na(value):
    """INEGI uses the literal string 'NA' for 'no aplica'. Map to None."""
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        if v == "" or v.upper() == "NA":
            return None
        try:
            return float(v.replace(",", ""))
        except ValueError:
            return None
    return float(value)


def _scan_years(ws):
    """Read row 5 of the cuadro: year labels mark the start of each 10-col block.

    Returns a list of dicts: {year, col_start, is_preliminary}.
    A 'P' suffix (e.g. '2024P') flags preliminary figures.
    """
    years = []
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(5, c).value
        if raw is None:
            continue
        label = str(raw).strip()
        is_prelim = label.endswith("P")
        digits = label[:-1] if is_prelim else label
        if digits.isdigit() and 2000 <= int(digits) <= 2035:
            years.append({"year": int(digits), "col_start": c,
                           "is_preliminary": is_prelim})
    return years


def extract_cuadro(cuadro_id: str) -> list:
    """Parse one CSCM O&U cuadro → list of long-format trade rows."""
    info = CUADROS[cuadro_id]
    wb = openpyxl.load_workbook(SRC / info["file"], data_only=True)
    ws = wb["Tabulado"]
    years = _scan_years(ws)
    if not years:
        sys.exit(f"  ✗ {cuadro_id}: no year columns found in row 5")

    rows = []
    current_general = None
    for r in range(INDENT_TOTAL, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        if raw is None or str(raw).strip() == "":
            break  # data block ends; notes / blank rows follow
        label = str(raw)
        indent = len(label) - len(label.lstrip(" "))
        name = label.strip()

        if indent == INDENT_TOTAL:
            area_level = "total"
            area_general = "Total cultura"
            area_especifica = None
        elif indent == INDENT_AREA_GENERAL:
            area_level = "area_general"
            current_general = name
            area_general = name
            area_especifica = None
        elif indent == INDENT_AREA_ESPECIFICA:
            area_level = "area_especifica"
            area_general = current_general
            area_especifica = name
        else:
            continue  # unexpected indent — skip defensively

        for y in years:
            fx = FX_MXN_PER_USD.get(y["year"]) if info["price_basis"] == "corriente" else None
            for flow, offset in (("importacion", OFFSET_IMPORTACION),
                                 ("exportacion", OFFSET_EXPORTACION)):
                val = _na(ws.cell(r, y["col_start"] + offset).value)
                usd = None
                if val is not None and fx:
                    usd = round(val / fx, 6)
                rows.append(dict(
                    year=y["year"],
                    area_level=area_level,
                    area_general=area_general,
                    area_especifica=area_especifica,
                    flow=flow,
                    price_basis=info["price_basis"],
                    value_mxn_million=val,
                    value_usd_million=usd,
                    fx_rate_mxn_per_usd=fx,
                    is_preliminary=y["is_preliminary"],
                    source_cuadro=cuadro_id,
                ))
    return rows


def write_parquet(df: pd.DataFrame, table_name: str, out_dir: Path) -> Path:
    """Write Parquet via DuckDB (no pyarrow). Returns the output path."""
    out_path = out_dir / f"{table_name}.parquet"
    con = duckdb.connect()
    con.register("df_data", df)
    con.execute(
        f"COPY df_data TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)"
    )
    size_kb = out_path.stat().st_size / 1024
    print(f"  ✓ {out_path.relative_to(REPO_ROOT)} — {len(df):,} rows, {size_kb:.1f} KB")
    return out_path


def write_meta(out_path: Path, source_files: list, description: str,
               extra: dict | None = None) -> None:
    """Write the <table>.meta.json sidecar (naming_conventions.md §Versioning)."""
    srcs = []
    for sf in source_files:
        p = Path(sf)
        h = hashlib.sha256(p.read_bytes()).hexdigest() if p.exists() else None
        srcs.append({"file": p.name, "sha256": h})
    meta = {
        "table": out_path.stem,
        "description": description,
        "source": "INEGI — Cuenta Satélite de la Cultura de México (CSCM), "
                   "Sistema de Cuentas Nacionales de México, año base 2018",
        "source_url": "https://www.inegi.org.mx/programas/cultura/2018/",
        "source_files": srcs,
        "fetch_date": "2026-05-22",
        "etl_script": "etl/inegi__csc_xlsx_to_parquet.py",
        "etl_run_date": str(date.today()),
        "licence": "INEGI — Términos de Libre Uso de la Información del INEGI",
    }
    if extra:
        meta.update(extra)
    meta_path = out_path.with_suffix(".meta.json")
    meta_path.write_text(json.dumps(meta, indent=2, ensure_ascii=False))
    print(f"  ✓ {meta_path.relative_to(REPO_ROOT)}")


def maybe_push(df: pd.DataFrame, schema: str, table: str) -> None:
    """Push to MotherDuck ONLY if MOTHERDUCK_TOKEN is set. Off by default."""
    token = os.environ.get("MOTHERDUCK_TOKEN")
    if not token:
        print(f"  · MotherDuck push skipped for atana.{schema}.{table} "
              f"(MOTHERDUCK_TOKEN not set)")
        return
    con = duckdb.connect(f"md:atana?motherduck_token={token}")
    con.execute(f"CREATE SCHEMA IF NOT EXISTS atana.{schema}")
    con.register("df_data", df)
    con.execute(
        f"CREATE OR REPLACE TABLE atana.{schema}.{table} AS SELECT * FROM df_data"
    )
    n = con.execute(f"SELECT COUNT(*) FROM atana.{schema}.{table}").fetchone()[0]
    print(f"  ✓ Synced atana.{schema}.{table} ({n:,} rows)")


def main() -> None:
    if not SRC.exists():
        sys.exit(f"Source folder not found: {SRC}\n"
                 f"Expected the unzipped CSCM tabulados there.")

    # ── 1. FX reference table ────────────────────────────────────────────────
    print("Building FX reference table...")
    fx_df = pd.DataFrame(
        [{"year": y, "fx_rate_mxn_per_usd": r, "source": FX_SOURCE}
         for y, r in sorted(FX_MXN_PER_USD.items())]
    )
    fx_path = write_parquet(fx_df, "fx_mxn_usd_annual", REF)
    write_meta(fx_path, [], "Annual-average MXN/USD exchange rate, used to "
               "derive the USD column of inegi.csc_comercio. NOT an INEGI "
               "figure — a documented convenience series.",
               extra={"source": FX_SOURCE,
                      "source_url": "https://datos.bancomundial.org/indicador/PA.NUS.FCRF?locations=MX"})

    # ── 2. csc_comercio — the trade module ───────────────────────────────────
    print("Extracting CSCM cultural-trade columns...")
    all_rows = []
    for cuadro_id in CUADROS:
        rows = extract_cuadro(cuadro_id)
        print(f"  · {cuadro_id}: {len(rows):,} rows parsed")
        all_rows.extend(rows)
    df = pd.DataFrame(all_rows)
    # Stable ordering for idempotent byte-identical output.
    df = df.sort_values(
        ["price_basis", "year", "flow", "area_level", "area_general",
         "area_especifica"],
        na_position="first").reset_index(drop=True)
    com_path = write_parquet(df, "csc_comercio", OUT)
    write_meta(com_path,
               [SRC / "CSCM_31.xlsx", SRC / "CSCM_75.xlsx"],
               "Cultural goods+services imports (Importaciones C.I.F.) and "
               "exports (Exportaciones) from the CSCM Cuadros de Oferta y "
               "Utilización, by functional area, 2008–2024, in million MXN "
               "(current and constant-2018 prices). USD column is ETL-derived.",
               extra={"grain": "year × area × flow × price_basis",
                      "currency": "MXN million (value_mxn_million); "
                                  "USD million is derived, see fx_mxn_usd_annual",
                      "note_no_bop_module": "The CSCM has no balance-of-payments "
                      "module; these figures are the import/export columns of "
                      "the supply-use tables (CSCM_31 corriente, CSCM_75 "
                      "constante 2018). No bilateral partner detail exists."})

    # ── 3. MotherDuck (gated, off by default) ────────────────────────────────
    maybe_push(fx_df, "inegi", "fx_mxn_usd_annual")
    maybe_push(df, "inegi", "csc_comercio")
    print("Done.")


if __name__ == "__main__":
    main()
