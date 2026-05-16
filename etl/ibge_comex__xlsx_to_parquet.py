"""Convert IBGE Comércio Exterior Cultural (Tabela 10.1–10.4) from xlsx to Parquet.

Source files live in the analysis repo at:
    ../10_comercio_exterior_de_bens_e_servicos_culturais/

Output (one Parquet file per source table, long format):
    raw/ibge_comex/tab_10_1.parquet   (bens, por capítulo NCM × ano)
    raw/ibge_comex/tab_10_2.parquet   (bens, % por capítulo)
    raw/ibge_comex/tab_10_3.parquet   (top 20 países parceiros)
    raw/ibge_comex/tab_10_4.parquet   (serviços audiovisuais BCB)

Also uploads to md:atana.ibge_comex.

Usage:
    export MOTHERDUCK_TOKEN="..."
    python etl/ibge_comex__xlsx_to_parquet.py
"""
import os
import sys
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "raw" / "ibge_comex"
OUT.mkdir(parents=True, exist_ok=True)

# Source: hard-coded path to the analysis project's xlsx files
SRC = Path(
    "/Users/joaoroque/Documents/Cultural production - book/"
    "Dados da Economia Cultural no Brasil/"
    "10_comercio_exterior_de_bens_e_servicos_culturais"
)
# Sandbox fallback path (used when running inside the agent workspace)
if not SRC.exists():
    SRC = Path(
        "/sessions/wizardly-admiring-feynman/mnt/"
        "Dados da Economia Cultural no Brasil/"
        "10_comercio_exterior_de_bens_e_servicos_culturais"
    )

YEARS = [str(y) for y in range(2014, 2025)]

# Chapter labels (5 chapters are 100% cultural; 13 are partially cultural)
CAPITULO_LABELS = {
    32: "Tintas/corantes",
    37: "Fotografia/cinema (100% cultural)",
    39: "Plástico",
    44: "Madeira",
    46: "Espartaria (100% cultural)",
    49: "Livros/gráfica (100% cultural)",
    67: "Penas/flores",
    69: "Cerâmica",
    71: "Joalheria",
    83: "Metais comuns",
    84: "Máquinas",
    85: "Eletro-AV",
    90: "Óptica",
    91: "Relojoaria",
    92: "Instrumentos musicais (100% cultural)",
    94: "Móveis/iluminação",
    95: "Brinquedos/jogos",
    97: "Objetos de arte (100% cultural)",
}


def extract_tab_10_1() -> pd.DataFrame:
    """Long format: year × capitulo_ncm × imp_total, imp_cult, exp_total, exp_cult, etc."""
    wb = openpyxl.load_workbook(
        SRC / "Tabela 10.1_Bens_Comercio_Exterior_Cultura_2014_24.xlsx",
        data_only=True,
    )
    rows = []
    for year in YEARS:
        ws = wb[year]
        # Row 6 = Brasil total
        rows.append(dict(
            year=int(year),
            capitulo_ncm=None,  # null = total
            capitulo_label="Total",
            imp_todos_produtos_brl_mi=ws.cell(6, 3).value,
            imp_cultural_brl_mi=ws.cell(6, 4).value,
            imp_pct_cultural=ws.cell(6, 5).value,
            exp_todos_produtos_brl_mi=ws.cell(6, 6).value,
            exp_cultural_brl_mi=ws.cell(6, 7).value,
            exp_pct_cultural=ws.cell(6, 8).value,
            is_pure_cultural=None,
        ))
        # Rows 8-25 = capítulos
        for r in range(8, 26):
            ch = ws.cell(r, 1).value
            if not isinstance(ch, int):
                continue
            rows.append(dict(
                year=int(year),
                capitulo_ncm=ch,
                capitulo_label=CAPITULO_LABELS.get(ch, f"Cap. {ch}"),
                imp_todos_produtos_brl_mi=ws.cell(r, 3).value,
                imp_cultural_brl_mi=ws.cell(r, 4).value,
                imp_pct_cultural=ws.cell(r, 5).value,
                exp_todos_produtos_brl_mi=ws.cell(r, 6).value,
                exp_cultural_brl_mi=ws.cell(r, 7).value,
                exp_pct_cultural=ws.cell(r, 8).value,
                is_pure_cultural=ch in {37, 46, 49, 92, 97},
            ))
    return pd.DataFrame(rows)


def extract_tab_10_2() -> pd.DataFrame:
    """Tab 10.2 — share of each capítulo within cultural total (Imp and Exp)."""
    wb = openpyxl.load_workbook(
        SRC / "Tabela 10.2_Bens_Comercio_Exterior_Cultura_2014_24.xlsx",
        data_only=True,
    )
    rows = []
    for year in YEARS:
        ws = wb[year]
        # Total row (6) = 100%, then capítulos in 8-25
        for r in range(8, 26):
            ch = ws.cell(r, 1).value
            if not isinstance(ch, int):
                continue
            rows.append(dict(
                year=int(year),
                capitulo_ncm=ch,
                capitulo_label=CAPITULO_LABELS.get(ch, f"Cap. {ch}"),
                imp_cultural_brl_mi=ws.cell(r, 3).value,
                imp_pct_no_total_cultural=ws.cell(r, 4).value,
                exp_cultural_brl_mi=ws.cell(r, 5).value,
                exp_pct_no_total_cultural=ws.cell(r, 6).value,
            ))
    return pd.DataFrame(rows)


def extract_tab_10_3() -> pd.DataFrame:
    """Top 20 partner countries × year × 4 flows."""
    wb = openpyxl.load_workbook(
        SRC / "Tabela 10.3_Vinte_Países_Imp_e_Exp_Bens_Cult_Total_2014_24.xlsx",
        data_only=True,
    )
    rows = []
    flow_columns = {
        "imp_total":   (1, 2),  # A, B
        "imp_cultural":(3, 4),  # C, D
        "exp_total":   (5, 6),  # E, F
        "exp_cultural":(7, 8),  # G, H
    }
    for year in YEARS:
        ws = wb[year]
        for flow_name, (name_col, pct_col) in flow_columns.items():
            rank = 1
            for r in range(5, 25):  # 20 rows
                country = ws.cell(r, name_col).value
                pct = ws.cell(r, pct_col).value
                if country and pct is not None:
                    rows.append(dict(
                        year=int(year),
                        flow=flow_name,
                        rank=rank,
                        country=country,
                        share_pct=pct,
                    ))
                    rank += 1
    return pd.DataFrame(rows)


def extract_tab_10_4() -> pd.DataFrame:
    """Audiovisual services balance — BCB Balance of Payments."""
    wb = openpyxl.load_workbook(
        SRC / "Tabela 10.4_Serviços_Audiovisuais_Comércio_Exterior_2014_24.xlsx",
        data_only=True,
    )
    rows = []
    for year in YEARS:
        ws = wb[year]
        rows.append(dict(
            year=int(year),
            serv_totais_saldo_brl_mi=ws.cell(3, 2).value,
            serv_totais_receita_brl_mi=ws.cell(4, 2).value,
            serv_totais_despesa_brl_mi=ws.cell(5, 2).value,
            av_saldo_brl_mi=ws.cell(7, 2).value,
            av_receita_brl_mi=ws.cell(8, 2).value,
            av_despesa_brl_mi=ws.cell(9, 2).value,
            av_receita_pct_servicos=ws.cell(11, 2).value,
            av_despesa_pct_servicos=ws.cell(12, 2).value,
        ))
    return pd.DataFrame(rows)


def load_and_push(df: pd.DataFrame, table_name: str) -> None:
    """Write Parquet via DuckDB (no pyarrow dependency) and push to MotherDuck."""
    out_path = OUT / f"{table_name}.parquet"
    token = os.environ.get("MOTHERDUCK_TOKEN")

    # Local DuckDB writes the Parquet without needing pyarrow
    local_con = duckdb.connect()
    local_con.register("df_data", df)
    local_con.execute(
        f"COPY df_data TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)"
    )
    size_kb = out_path.stat().st_size / 1024
    print(f"  ✓ {out_path.name} — {len(df):,} rows, {size_kb:.1f} KB")

    if token:
        con = duckdb.connect(f"md:atana?motherduck_token={token}")
        con.execute("CREATE SCHEMA IF NOT EXISTS atana.ibge_comex")
        con.register("df_data", df)
        con.execute(
            f"CREATE OR REPLACE TABLE atana.ibge_comex.{table_name} AS SELECT * FROM df_data"
        )
        n = con.execute(f"SELECT COUNT(*) FROM atana.ibge_comex.{table_name}").fetchone()[0]
        print(f"    Synced atana.ibge_comex.{table_name} ({n:,} rows)")


def main() -> None:
    if not SRC.exists():
        sys.exit(f"Source folder not found: {SRC}")
    print("Extracting IBGE Comex tables...")
    for name, fn in [
        ("tab_10_1", extract_tab_10_1),
        ("tab_10_2", extract_tab_10_2),
        ("tab_10_3", extract_tab_10_3),
        ("tab_10_4", extract_tab_10_4),
    ]:
        df = fn()
        load_and_push(df, name)


if __name__ == "__main__":
    main()
