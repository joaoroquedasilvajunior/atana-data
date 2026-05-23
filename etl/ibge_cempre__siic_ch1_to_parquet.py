"""SIIC Chapter 1 — Atividades formalmente constituídas → Parquet.

Phase 4a of the Atana Data expansion. The firm-structure complement to chapter 2:
IBGE's CEMPRE (Cadastro Central de Empresas) plus the company-demography and
public-register statistics give local units, employment, wages and firm
demography for the cultural sector — the registered-business view of the FCS
'Cultural and creative goods manufacturing' domain (and of the cultural economy
more broadly).

WHAT THIS INGESTS
-----------------
Source (already in the workspace — SIIC 'Informações Culturais', 2024 edition):
    1_atividades_formalmente_constituidas.xlsx
        24 sheets. 'Quadro 1.1' is a structure legend (skipped — not data).
        The 23 data sheets are heterogeneous:
          1.1.x  CEMPRE statistics (all legal natures)        — ref. 2022
          1.2.x  company demography (births/deaths/survival)  — ref. 2015-2022
          1.3.x  public-register statistics (legal nature 213-5 …)
        '.a' sheets are provisional support tables.

The 23 sheets do not share one layout (some are single-year, some multi-block,
column 1 is a year in some and a label in others). Rather than 23 bespoke
parsers, this ETL preserves each sheet faithfully — the `ibge_pnadc` precedent:
one row per non-empty spreadsheet row, original cells kept as c01, c02, …, over
the sheet's real used range. Column and row meaning is documented in
docs/methodology/ibge_cempre_siic_ch1.md.

OUTPUT, one Parquet per data sheet:
    raw/ibge_cempre/tab_1_1_1.parquet … tab_1_3_4.parquet (+ .meta.json)
    grain: one row per source spreadsheet row (row_index preserves order)

Idempotent. openpyxl read → list-of-dicts → DuckDB COPY to Parquet (no pyarrow).
MotherDuck push gated behind a token (skipped when ATANA_ETL_SKIP_PUSH is set).
Schema: atana.ibge_cempre.

Usage:
    python etl/ibge_cempre__siic_ch1_to_parquet.py
"""
import glob
import hashlib
import json
import os
from datetime import date
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "raw" / "ibge_cempre"
OUT.mkdir(parents=True, exist_ok=True)

_WS_CANDIDATES = [
    "/Users/joaoroque/Documents/Cultural production - book/"
    "Dados da Economia Cultural no Brasil",
    *sorted(glob.glob("/sessions/*/mnt/Dados da Economia Cultural no Brasil")),
]
WORKSPACE = next((Path(p) for p in _WS_CANDIDATES if Path(p).exists()), None)
SOURCE_FILE = (WORKSPACE / "1_atividades_formalmente_constituidas.xlsx"
               if WORKSPACE else None)

SKIP_SHEETS = {"Quadro 1.1"}            # structure legend, not statistical data
SKIP_PREFIXES = ("Tabela", "Fonte", "Nota", "Quadro")


def extract_sheet(ws, table_id: str):
    """One sheet → (DataFrame faithful wide, title, used_col_count)."""
    rows = list(ws.iter_rows(values_only=True))
    title = None
    if rows:
        title = next((str(v).strip() for v in rows[0] if v is not None), None)

    max_col = 0
    for row in rows:
        for i in range(len(row) - 1, -1, -1):
            if row[i] is not None:
                max_col = max(max_col, i + 1)
                break
    if max_col == 0:
        return pd.DataFrame(), title, 0

    records = []
    for r_idx, row in enumerate(rows, start=1):
        cells = list(row[:max_col])
        if all(v is None for v in cells):
            continue
        first = next((v for v in cells if v is not None), None)
        if isinstance(first, str) and first.strip().startswith(SKIP_PREFIXES):
            continue
        rec = {"table_id": table_id, "row_index": r_idx}
        for i, v in enumerate(cells, start=1):
            rec[f"c{i:02d}"] = v
        records.append(rec)
    return pd.DataFrame(records), title, max_col


def write_parquet(df: pd.DataFrame, table_name: str) -> Path:
    out_path = OUT / f"{table_name}.parquet"
    con = duckdb.connect()
    con.register("df_data", df)
    con.execute(f"COPY df_data TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    size_kb = out_path.stat().st_size / 1024
    print(f"  ✓ {out_path.relative_to(REPO_ROOT)} — {len(df):,} rows, {size_kb:.1f} KB")
    return out_path


def write_meta(out_path: Path, table_id: str, title: str, ncol: int) -> None:
    h = (hashlib.sha256(SOURCE_FILE.read_bytes()).hexdigest()
         if SOURCE_FILE and SOURCE_FILE.exists() else None)
    meta = {
        "table": out_path.stem,
        "description": f"SIIC chapter 1, sheet {table_id} — "
                       f"{title or 'CEMPRE / company-demography table'}. "
                       "Faithful wide image of the IBGE sheet (cells kept as "
                       "c01…); see the methodology note for column meaning.",
        "source": "IBGE — Sistema de Informações e Indicadores Culturais "
                   "(SIIC), 'Informações Culturais' 2024 edition, chapter 1 "
                   "'Atividades formalmente constituídas' (CEMPRE + company "
                   "demography + public-register statistics)",
        "source_url": "https://www.ibge.gov.br/estatisticas/sociais/cultura.html",
        "source_files": [{"file": SOURCE_FILE.name if SOURCE_FILE else None,
                          "sha256": h}],
        "fetch_date": "2026-05-23",
        "etl_script": "etl/ibge_cempre__siic_ch1_to_parquet.py",
        "etl_run_date": str(date.today()),
        "licence": "open data — IBGE official statistics",
        "grain": "one row per source spreadsheet row (row_index preserves order)",
        "used_columns": ncol,
        "notes": "Faithful preservation: header rows and data rows are both "
                 "kept (row_index disambiguates). Column 1 (c01) is a year in "
                 "some tables and a label in others — see the methodology note.",
    }
    out_path.with_suffix(".meta.json").write_text(
        json.dumps(meta, indent=2, ensure_ascii=False))
    print(f"  ✓ {out_path.with_suffix('.meta.json').relative_to(REPO_ROOT)}")


def maybe_push(df: pd.DataFrame, schema: str, table: str) -> None:
    """Push to MotherDuck if a valid JWT token is available.

    Skipped entirely when ATANA_ETL_SKIP_PUSH is set — used for build-only /
    sandbox runs where the Parquet is produced but the sync stays with João.
    """
    if os.environ.get("ATANA_ETL_SKIP_PUSH"):
        print(f"  · push skipped for atana.{schema}.{table} (ATANA_ETL_SKIP_PUSH)")
        return

    def _jwt(t) -> str:
        t = (t or "").strip()
        return t if (t.startswith("eyJ") and t.count(".") == 2) else ""
    token = _jwt(os.environ.get("MOTHERDUCK_TOKEN"))
    if not token:
        tf = REPO_ROOT / ".motherduck_token"
        token = _jwt(tf.read_text()) if tf.exists() else ""
    if not token:
        print(f"  · MotherDuck push skipped for atana.{schema}.{table} — no "
              f"valid token. Put a JWT in {REPO_ROOT}/.motherduck_token, re-run.")
        return
    con = duckdb.connect(f"md:atana?motherduck_token={token}")
    con.execute(f"CREATE SCHEMA IF NOT EXISTS atana.{schema}")
    con.register("df_data", df)
    con.execute(
        f"CREATE OR REPLACE TABLE atana.{schema}.{table} AS SELECT * FROM df_data")
    n = con.execute(f"SELECT COUNT(*) FROM atana.{schema}.{table}").fetchone()[0]
    print(f"  ✓ Synced atana.{schema}.{table} ({n:,} rows)")


def main() -> None:
    if not SOURCE_FILE or not SOURCE_FILE.exists():
        raise SystemExit(f"Source file not found (workspace: {WORKSPACE})")
    print(f"Reading {SOURCE_FILE.name}")
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            print(f"  · skipped legend sheet '{sheet_name}'")
            continue
        table_id = sheet_name.strip()
        df, title, ncol = extract_sheet(wb[sheet_name], table_id)
        if df.empty:
            print(f"  ⚠ sheet '{sheet_name}' empty — skipped")
            continue
        table_name = "tab_" + table_id.replace(".", "_").lower()
        out_path = write_parquet(df, table_name)
        write_meta(out_path, table_id, title, ncol)
        maybe_push(df, "ibge_cempre", table_name)
    print("Done.")


if __name__ == "__main__":
    main()
