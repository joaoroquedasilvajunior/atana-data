"""INPI — Tabelas Completas dos Indicadores de Propriedade Industrial → Parquet.

Phase 4c.2 of the Atana Data expansion. Enriches the FCS *Intellectual property*
domain (already reached by 4c.1 BCB) with the **cultural-IP stock** — the
registration counts a balance-of-payments flow cannot show.

WHAT THIS INGESTS
-----------------
Source (downloaded by João into raw/inpi/ — the INPI "Tabelas Completas dos
Indicadores de PI" page, anuário edition):
    raw/inpi/indicadores_anuario_2024.zip
        8 workbooks, one per IP type. This ETL takes the FOUR cultural ones:
            Indicadores_PRG_Anuario_2024.xlsx     — computer programs
            Indicadores_DI_Anuario_2024.xlsx      — industrial designs
            Indicadores_IG_Anuario_2024.xlsx      — geographical indications
            Indicadores_Marcas_Anuario_2024.xlsx  — trademarks (by Nice class)
        and SKIPS patents (PTN), technology contracts (CTT), IC topographies
        (TCI) and the gender cut (PI_Genero) — not cultural IP.

Each workbook holds one indicator table per sheet; tables are 2000-2024 annual
series. The 2024 edition carries the full history, so it is ingested alone;
the older editions (2018-2020, 2023) remain in raw/inpi/ for cross-checking.

The sheets are not one shape (year-in-rows series; class-in-rows / year-in-
columns breakdowns). Per the `ibge_cempre` precedent the ETL preserves each
sheet faithfully — one Parquet per sheet, original cells kept as c01, c02, ….

⚠️ ALL-IP, FILTERED AT THE FILE LEVEL. The cultural cut here is "the four
cultural IP types". Computer programs and industrial designs are creative by
nature; for trademarks the cultural slice is a Nice-class filter (default tight
= classes 41 + 16; wide = + 9 + 28) applied downstream on the `mrc_*_classe*`
tables — the class dimension is preserved, not pre-filtered. See the
methodology note.

OUTPUT, one Parquet per source sheet (~68 tables):
    raw/inpi/<iptype>_<sheet>.parquet  (+ .meta.json)   iptype ∈ {prg,di,ig,mrc}

Idempotent. zip→openpyxl→list-of-dicts→DuckDB COPY to Parquet (no pyarrow).
MotherDuck push gated behind a token (skipped when ATANA_ETL_SKIP_PUSH is set).
Schema: atana.inpi.

Usage:
    python etl/inpi__indicadores_to_parquet.py
"""
import hashlib
import io
import json
import os
import re
import unicodedata
import zipfile
from datetime import date
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "raw" / "inpi"
OUT.mkdir(parents=True, exist_ok=True)

_ZIP_NAME = "indicadores_anuario_2024.zip"
SOURCE_ZIP = next(
    (p for p in (OUT / "_source" / _ZIP_NAME, OUT / _ZIP_NAME) if p.exists()),
    OUT / _ZIP_NAME)  # the source zips may be kept in raw/inpi/ or raw/inpi/_source/

# cultural IP types — workbook-filename keyword → iptype code
CULTURAL = {"_PRG_": "prg", "_DI_": "di", "_IG_": "ig", "_Marcas_": "mrc"}
SKIP_SHEETS = {"sumario", "sumário"}
SKIP_ROW_PREFIXES = ("Anuário", "Fonte", "Nota")


def slug(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s).strip().lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "_", s).strip("_")


def extract_sheet(ws, iptype: str, sheet: str) -> pd.DataFrame:
    """One sheet → faithful wide rows (the ibge_cempre preservation pattern)."""
    rows = list(ws.iter_rows(values_only=True))
    max_col = 0
    for row in rows:
        for i in range(len(row) - 1, -1, -1):
            if row[i] is not None:
                max_col = max(max_col, i + 1)
                break
    if max_col == 0:
        return pd.DataFrame()
    records = []
    for r_idx, row in enumerate(rows, start=1):
        cells = list(row[:max_col])
        if all(v is None for v in cells):
            continue
        first = next((v for v in cells if v is not None), None)
        if isinstance(first, str) and first.strip().startswith(SKIP_ROW_PREFIXES):
            continue
        rec = {"iptype": iptype, "source_sheet": sheet, "row_index": r_idx}
        for i, v in enumerate(cells, start=1):
            # store every cell as text — uniform VARCHAR columns, faithful to
            # the source; downstream casts as needed. Mixing typed and string
            # cells in one column otherwise trips DuckDB's type inference.
            rec[f"c{i:02d}"] = None if v is None else str(v)
        records.append(rec)
    return pd.DataFrame(records)


def write_parquet(df: pd.DataFrame, table_name: str) -> Path:
    out_path = OUT / f"{table_name}.parquet"
    con = duckdb.connect()
    con.register("df_data", df)
    con.execute(f"COPY df_data TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    size_kb = out_path.stat().st_size / 1024
    print(f"  ✓ {out_path.name} — {len(df):,} rows, {size_kb:.1f} KB")
    return out_path


def write_meta(out_path: Path, iptype: str, workbook: str, sheet: str,
               ncol: int) -> None:
    h = (hashlib.sha256(SOURCE_ZIP.read_bytes()).hexdigest()
         if SOURCE_ZIP.exists() else None)
    meta = {
        "table": out_path.stem,
        "schema": "inpi",
        "ip_type": iptype,
        "description": f"INPI Indicadores de PI — {iptype.upper()} — sheet "
                       f"'{sheet}'. Annual series 2000-2024, anuário edition. "
                       "Faithful wide image of the source sheet (cells kept as "
                       "c01…); see docs/methodology/inpi_indicadores.md.",
        "source": "INPI — Instituto Nacional da Propriedade Industrial, "
                  "Tabelas Completas dos Indicadores de Propriedade Industrial "
                  "(Anuário Estatístico), 2024 edition",
        "source_url": "https://www.gov.br/inpi/pt-br/inpi-data/"
                      "dados-e-series-temporais/"
                      "tabelas-completas-dos-indicadores-de-pi",
        "source_files": [{"file": SOURCE_ZIP.name, "member": workbook,
                          "sha256": h}],
        "fetch_date": "2026-05-23",
        "etl_script": "etl/inpi__indicadores_to_parquet.py",
        "etl_run_date": str(date.today()),
        "licence": "open data — INPI official statistics",
        "grain": "faithful wide image — one row per source spreadsheet row",
        "used_columns": ncol,
        "notes": "Cultural IP cut at the file level (prg/di/ig/mrc only). "
                 "Trademark cultural slice = Nice classes 41+16 (tight) or "
                 "+9+28 (wide), applied downstream on the mrc class tables — "
                 "the class dimension is preserved here, not pre-filtered.",
    }
    out_path.with_suffix(".meta.json").write_text(
        json.dumps(meta, indent=2, ensure_ascii=False))


def maybe_push(df: pd.DataFrame, schema: str, table: str) -> None:
    """Push to MotherDuck if a valid JWT token is available.

    Skipped entirely when ATANA_ETL_SKIP_PUSH is set — for build-only / sandbox
    runs where the Parquet is produced but the sync stays with João.
    """
    if os.environ.get("ATANA_ETL_SKIP_PUSH"):
        return

    def _jwt(t) -> str:
        t = (t or "").strip()
        return t if (t.startswith("eyJ") and t.count(".") == 2) else ""
    token = _jwt(os.environ.get("MOTHERDUCK_TOKEN"))
    if not token:
        tf = REPO_ROOT / ".motherduck_token"
        token = _jwt(tf.read_text()) if tf.exists() else ""
    if not token:
        print(f"  · MotherDuck push skipped for atana.{schema}.{table} — no "
              f"valid token.")
        return
    con = duckdb.connect(f"md:atana?motherduck_token={token}")
    con.execute(f"CREATE SCHEMA IF NOT EXISTS atana.{schema}")
    con.register("df_data", df)
    con.execute(
        f"CREATE OR REPLACE TABLE atana.{schema}.{table} AS SELECT * FROM df_data")
    print(f"  ✓ Synced atana.{schema}.{table}")


def main() -> None:
    if not SOURCE_ZIP.exists():
        raise SystemExit(f"Source zip not found: {SOURCE_ZIP}")
    print(f"Reading {SOURCE_ZIP.name}")
    zf = zipfile.ZipFile(SOURCE_ZIP)
    n_tables = 0
    for member in sorted(zf.namelist()):
        iptype = next((v for k, v in CULTURAL.items() if k in member), None)
        if not iptype or not member.lower().endswith((".xlsx", ".xls")):
            continue
        print(f"\n{member}  →  iptype '{iptype}'")
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(member)), data_only=True)
        for sheet in wb.sheetnames:
            if sheet.strip().lower() in SKIP_SHEETS:
                continue
            df = extract_sheet(wb[sheet], iptype, sheet)
            if df.empty:
                continue
            table = f"{iptype}_{slug(sheet)}"
            ncol = sum(1 for c in df.columns if c.startswith("c"))
            out_path = write_parquet(df, table)
            write_meta(out_path, iptype, member, sheet, ncol)
            maybe_push(df, "inpi", table)
            n_tables += 1
    print(f"\nDone — {n_tables} tables written to raw/inpi/.")


if __name__ == "__main__":
    main()
