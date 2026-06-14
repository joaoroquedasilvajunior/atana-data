"""LPG (Lei Paulo Gustavo, LC 195/2022) → schema atana.lpg.

Phase LPG ingest (2026-06-14). Four CSVs + two XLSX from MinC dados.cultura.gov.br
(DFD/SEFIC, atualização mensal) → four curated Parquet tables.

Idempotent: stable sort + DuckDB COPY (ZSTD), byte-identical reruns.
MotherDuck sync is a manual checkpoint (NEW schema).

SOURCE FILES (place in raw/lpg/_source/):
  adesaoestadoslpg.csv               — 27 rows, 3 cols
  adesaomunicipioslpg.csv            — 5,569 rows, 4 cols
  execucaofinanceiraestadoslpg.csv   — 54 rows, 8 cols (no IBGE)
  execucaofinanceiramunicipioslpg.csv — 10,930 rows, 10 cols (has IBGE)
  extratobancariolpg.xlsx            — Sheet1, 10,968 rows, 10 cols
  relatoriogestaolpg.xlsx            — Sheet1, 18,181 rows, 6 cols

KEY FINDING (the corrected mental model):
  Meta do Plano splits the total R$ 3.86 bi LPG into:
    - 'Audiovisual'  → R$ 2.80 bi (AV-emergency, the AV-targeted portion)
    - 'Outras Áreas' → R$ 1.07 bi (post-pandemic generalist cultural support)
  Both flow through the same MinC ente-federado pipe. There is NO separate
  FSA/ANCINE route in this dataset — LPG-AV money is direct-to-ente, just
  flagged by meta. (The earlier "R$ 2.79 bi via FSA" mental model was wrong.)

HARMONIZATION:
  - tipo_ente ∈ {'Estado', 'Município'} distinguishes the two universes
  - cod_ibge: 2-digit IBGE state code for estados, 7-digit IBGE municipal
    code for municípios — uniform numeric column
  - Estado-name → IBGE-2-digit lookup hard-coded below (27 entries)
  - Date format in CSVs: MM/DD/YYYY → ISO YYYY-MM-DD

OUTPUT TABLES:
  atana.lpg.adesao_entes       — 5,595 rows
  atana.lpg.execucao_financeira — 10,984 rows
  atana.lpg.extratos_bancarios  — 10,968 rows
  atana.lpg.relatorio_gestao    — 18,181 rows
"""
import re
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "raw" / "lpg" / "_source"
OUT = REPO / "curated" / "lpg"
OUT.mkdir(parents=True, exist_ok=True)


# ─── Brazilian state-name → IBGE 2-digit code lookup ──────────────────────
UF_NAME_TO_CODE = {
    "ACRE": 12, "ALAGOAS": 27, "AMAPÁ": 16, "AMAPA": 16,
    "AMAZONAS": 13, "BAHIA": 29, "CEARÁ": 23, "CEARA": 23,
    "DISTRITO FEDERAL": 53, "ESPÍRITO SANTO": 32, "ESPIRITO SANTO": 32,
    "GOIÁS": 52, "GOIAS": 52, "MARANHÃO": 21, "MARANHAO": 21,
    "MATO GROSSO": 51, "MATO GROSSO DO SUL": 50, "MINAS GERAIS": 31,
    "PARÁ": 15, "PARA": 15, "PARAÍBA": 25, "PARAIBA": 25,
    "PARANÁ": 41, "PARANA": 41, "PERNAMBUCO": 26, "PIAUÍ": 22, "PIAUI": 22,
    "RIO DE JANEIRO": 33, "RIO GRANDE DO NORTE": 24, "RIO GRANDE DO SUL": 43,
    "RONDÔNIA": 11, "RONDONIA": 11, "RORAIMA": 14,
    "SANTA CATARINA": 42, "SÃO PAULO": 35, "SAO PAULO": 35,
    "SERGIPE": 28, "TOCANTINS": 17,
}

UF_ABBREV_TO_CODE = {
    "AC": 12, "AL": 27, "AP": 16, "AM": 13, "BA": 29, "CE": 23, "DF": 53,
    "ES": 32, "GO": 52, "MA": 21, "MT": 51, "MS": 50, "MG": 31, "PA": 15,
    "PB": 25, "PR": 41, "PE": 26, "PI": 22, "RJ": 33, "RN": 24, "RS": 43,
    "RO": 11, "RR": 14, "SC": 42, "SP": 35, "SE": 28, "TO": 17,
}

UF_CODE_TO_ABBREV = {v: k for k, v in UF_ABBREV_TO_CODE.items()}


# ─── helpers ──────────────────────────────────────────────────────────────
def f(x):
    if x is None or x == "":
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def s(x):
    return str(x).strip() if x is not None and x != "" else None


def estado_code(name):
    if name is None:
        return None
    return UF_NAME_TO_CODE.get(str(name).strip().upper())


def estado_uf(name):
    code = estado_code(name)
    return UF_CODE_TO_ABBREV.get(code) if code else None


def to_iso_date(x):
    """MM/DD/YYYY → YYYY-MM-DD; also accepts ISO already."""
    if x is None:
        return None
    t = str(x).strip()
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", t)
    if m:
        return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", t)
    return m.group(0) if m else None


def write(df, table, sort_cols):
    """Write df to curated/lpg/<table>.parquet, stably sorted."""
    con = duckdb.connect()
    con.register("df", df)
    con.execute("CREATE TABLE t AS SELECT * FROM df")
    order = ", ".join(f'"{c}"' for c in sort_cols)
    con.execute(f"CREATE TABLE s AS SELECT * FROM t ORDER BY {order} NULLS LAST")
    p = OUT / f"{table}.parquet"
    con.execute(f"COPY s TO '{p}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    n = con.execute("SELECT count(*) FROM s").fetchone()[0]
    con.close()
    print(f"  ✅ {table}: {n:,} rows → {p.relative_to(REPO)}")


# ─── 1. adesao_entes (estados ∪ municípios) ───────────────────────────────
def etl_adesao():
    rows = []

    # Estados
    df_e = pd.read_csv(SRC / "adesaoestadoslpg.csv")
    for _, r in df_e.iterrows():
        rows.append({
            "tipo_ente": "Estado",
            "cod_ibge": estado_code(r["Estado"]),
            "uf": estado_uf(r["Estado"]),
            "ente": s(r["Estado"]),
            "situacao_plano": s(r["Situação do Plano"]),
            "valor_disponivel_brl": f(r["Valor Disponível"]),
        })

    # Municípios — no IBGE in adesao file, so we cannot infer cod_ibge from
    # the adesao CSV alone. Leave cod_ibge NULL here; the execucao CSV has
    # IBGE codes and joins on (uf, ente).
    df_m = pd.read_csv(SRC / "adesaomunicipioslpg.csv")
    for _, r in df_m.iterrows():
        rows.append({
            "tipo_ente": "Município",
            "cod_ibge": None,
            "uf": s(r["UF"]),
            "ente": s(r["Município"]),
            "situacao_plano": s(r["Situação do Plano"]),
            "valor_disponivel_brl": f(r["Valor Disponível"]),
        })

    df = pd.DataFrame(rows)
    write(df, "adesao_entes", ["tipo_ente", "uf", "ente"])


# ─── 2. execucao_financeira (estados ∪ municípios; both metas preserved) ──
def etl_execucao():
    rows = []

    # Estados — no IBGE municipal code, map name → 2-digit code
    df_e = pd.read_csv(SRC / "execucaofinanceiraestadoslpg.csv")
    for _, r in df_e.iterrows():
        rows.append({
            "tipo_ente": "Estado",
            "cod_ibge": estado_code(r["Estado"]),
            "uf": estado_uf(r["Estado"]),
            "ente": s(r["Estado"]),
            "meta_plano": s(r["Meta do Plano"]),
            "data_pagamento": to_iso_date(r["Data Pagamento"]),
            "valor_recebido_brl": f(r["Valor Recebido"]),
            "rendimento_brl": f(r["Rendimento*"]),
            "saldo_brl": f(r["Saldo em conta"]),
            "valor_utilizado_brl": f(r["Valor Utilizado (R$)"]),
            "pct_utilizado_decimal": f(r["Valor Utilizado (%)"]),
        })

    # Municípios — IBGE code is present (column 'IBGE'); paid column named
    # 'Valor Transferido' instead of 'Valor Recebido' (same meaning)
    df_m = pd.read_csv(SRC / "execucaofinanceiramunicipioslpg.csv")
    for _, r in df_m.iterrows():
        rows.append({
            "tipo_ente": "Município",
            "cod_ibge": int(r["IBGE"]) if pd.notna(r["IBGE"]) else None,
            "uf": s(r["UF"]),
            "ente": s(r["Município"]),
            "meta_plano": s(r["Meta do Plano"]),
            "data_pagamento": to_iso_date(r["Data Pgto"]),
            "valor_recebido_brl": f(r["Valor Transferido"]),
            "rendimento_brl": f(r["Rendimento*"]),
            "saldo_brl": f(r["Saldo em conta"]),
            "valor_utilizado_brl": f(r["Valor Utilizado (R$)"]),
            "pct_utilizado_decimal": f(r.get("Valor Utilizado (%)**")),
        })

    df = pd.DataFrame(rows)
    write(df, "execucao_financeira",
          ["tipo_ente", "uf", "ente", "meta_plano"])


# ─── 3. extratos_bancarios (xlsx Sheet1, 10 cols) ─────────────────────────
def etl_extratos():
    wb = openpyxl.load_workbook(
        SRC / "extratobancariolpg.xlsx", read_only=True, data_only=True
    )
    ws = wb["Sheet1"]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue  # header
        rows.append({
            "nome_programa": s(r[0]),
            "codigo_plano_acao": s(r[1]),
            "uf_recebedor": s(r[2]),
            "municipio_recebedor": s(r[3]),
            "cnpj_solicitante": re.sub(r"\D", "", str(r[4])) if r[4] else None,
            "nome_solicitante": s(r[5]),
            "banco": s(r[6]),
            "agencia": s(r[7]),
            "conta": s(r[8]),
            "saldo_em_conta_brl": f(r[9]),
        })
    df = pd.DataFrame(rows)
    write(df, "extratos_bancarios",
          ["uf_recebedor", "municipio_recebedor", "codigo_plano_acao", "conta"])


# ─── 4. relatorio_gestao (xlsx Sheet1, 6 cols — narrative + execução %) ───
def etl_relatorio():
    wb = openpyxl.load_workbook(
        SRC / "relatoriogestaolpg.xlsx", read_only=True, data_only=True
    )
    ws = wb["Sheet1"]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        rows.append({
            "codigo_plano_acao": s(r[0]),
            "link_plano_acao": s(r[1]),
            "situacao_relatorio": s(r[2]),
            "meta_artigo": s(r[3]),
            "acao_descricao": s(r[4]),
            "execucao_fisica_decimal": f(r[5]),
        })
    df = pd.DataFrame(rows)
    write(df, "relatorio_gestao",
          ["codigo_plano_acao", "meta_artigo"])


# ─── main ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if not SRC.exists():
        raise SystemExit(
            f"❌ Source dir missing: {SRC}\n"
            "Place the 6 LPG files (4 CSVs + 2 XLSX) there, then rerun."
        )

    print(f"=== LPG ingest — reading from {SRC.relative_to(REPO)} ===")
    etl_adesao()
    etl_execucao()
    etl_extratos()
    etl_relatorio()
    print()
    print("Done. 4 tables in curated/lpg/. Validate next:")
    print("  python3 -c \"import duckdb; con=duckdb.connect();\\")
    print("    [print(f'{t:24s} {con.execute(chr(34)+f\\\"SELECT COUNT(*) FROM \\u0027curated/lpg/{t}.parquet\\u0027\\\"+chr(34)).fetchone()[0]:>7,}') \\")
    print("     for t in ['adesao_entes','execucao_financeira','extratos_bancarios','relatorio_gestao']]\"")
