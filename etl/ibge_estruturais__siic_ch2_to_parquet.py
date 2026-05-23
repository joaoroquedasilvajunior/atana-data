"""SIIC Chapter 2 — Pesquisas estruturais em empresas → Parquet.

Phase 4a of the Atana Data expansion. Partially closes the FCS transversal
domain 'Cultural and creative goods manufacturing': IBGE's structural business
surveys (PIA industrial / PAS services / PAC commerce) give production-account
measures — gross output, intermediate consumption, value added — that a
foreign-trade module cannot see.

WHAT THIS INGESTS
-----------------
Source (already in the workspace — SIIC 'Informações Culturais', 2024 edition):
    2_pesquisas_estruturais_em_empresas.xlsx
        8 sheets '2.1'..'2.8' — one structural-survey variable each, for the
        total economy and the cultural sector broken down by cultural domain
        and activity (CNAE class). Reference years 2013 + 2019-2023, every
        value paired with an IBGE CV reliability code (A/B/C/D/E/Z).

        2.1 Número de empresas              2.5 Custos e despesas
        2.2 Pessoal ocupado                 2.6 Valor bruto da produção
        2.3 Salários e remunerações         2.7 Consumo intermediário
        2.4 Receita líquida                 2.8 Valor adicionado

Each sheet: row 3 carries the variable caption; a year row holds the reference
years; the next row splits each year into 'Total' and 'CV'; data rows follow,
column 1 = the domain/activity label.

OUTPUT (long format), one Parquet per source sheet:
    raw/ibge_estruturais/tab_2_1.parquet .. tab_2_8.parquet (+ .meta.json)
    grain: table_id × variable × row_label × year → value, cv

Idempotent. openpyxl read → list-of-dicts → DuckDB COPY to Parquet (no pyarrow).
MotherDuck push gated behind a token. Schema: atana.ibge_estruturais.

Usage:
    python etl/ibge_estruturais__siic_ch2_to_parquet.py
"""
import glob
import hashlib
import json
import os
from datetime import date
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "raw" / "ibge_estruturais"
OUT.mkdir(parents=True, exist_ok=True)

# The SIIC xlsx lives in the analyst workspace, not in the repo. Resolve it
# whether the ETL runs on João's Mac or in a sandbox mount.
_WS_CANDIDATES = [
    "/Users/joaoroque/Documents/Cultural production - book/"
    "Dados da Economia Cultural no Brasil",
    *sorted(glob.glob("/sessions/*/mnt/Dados da Economia Cultural no Brasil")),
]
WORKSPACE = next((Path(p) for p in _WS_CANDIDATES if Path(p).exists()), None)
SOURCE_FILE = (WORKSPACE / "2_pesquisas_estruturais_em_empresas.xlsx"
               if WORKSPACE else None)

SHEETS = ["2.1", "2.2", "2.3", "2.4", "2.5", "2.6", "2.7", "2.8"]
SKIP_PREFIXES = ("Fonte", "Nota", "(*)", "(1)", "(2)", "(3)", "(4)", "(5)")


def _num(value):
    """Coerce a cell to float; blanks / dashes / non-numeric → None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s in ("", "-", "--", "...", "..", "n.d.", "X", "x"):
        return None
    try:
        return float(s.replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def _find_year_row(ws):
    """The header row that carries the reference years (>= 3 integer years)."""
    for r in range(1, 12):
        ints = [ws.cell(r, c).value for c in range(2, ws.max_column + 1)]
        ints = [v for v in ints
                if isinstance(v, (int, float)) and 2000 <= int(v) <= 2035]
        if len(ints) >= 3:
            return r
    return 4


def extract_sheet(wb, table_id: str) -> pd.DataFrame:
    """One sheet '2.X' → long-format rows (table_id, variable, label, year)."""
    ws = wb[table_id]
    year_row = _find_year_row(ws)

    caption = ws.cell(year_row - 1, 2).value
    variable = str(caption).strip() if caption else None

    # year → value column; the CV code sits in the next column
    year_cols = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(year_row, c).value
        if isinstance(v, (int, float)) and 2000 <= int(v) <= 2035:
            year_cols.append((int(v), c))

    rows = []
    for r in range(year_row + 2, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not isinstance(label, str) or not label.strip():
            continue
        label = label.strip()
        if label.startswith(SKIP_PREFIXES):
            break  # the footnote block — nothing tabular after it
        for year, c in year_cols:
            cv = ws.cell(r, c + 1).value
            rows.append(dict(
                table_id=table_id,
                variable=variable,
                row_index=r,
                row_label=label,
                year=year,
                value=_num(ws.cell(r, c).value),
                cv=(cv.strip() if isinstance(cv, str) and cv.strip() else None),
            ))
    return pd.DataFrame(rows)


def write_parquet(df: pd.DataFrame, table_name: str) -> Path:
    out_path = OUT / f"{table_name}.parquet"
    con = duckdb.connect()
    con.register("df_data", df)
    con.execute(f"COPY df_data TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    size_kb = out_path.stat().st_size / 1024
    print(f"  ✓ {out_path.relative_to(REPO_ROOT)} — {len(df):,} rows, {size_kb:.1f} KB")
    return out_path


def write_meta(out_path: Path, table_id: str, variable: str) -> None:
    h = (hashlib.sha256(SOURCE_FILE.read_bytes()).hexdigest()
         if SOURCE_FILE and SOURCE_FILE.exists() else None)
    meta = {
        "table": out_path.stem,
        "description": f"SIIC chapter 2, table {table_id} — {variable}. "
                       "IBGE structural business surveys (PIA/PAS/PAC), total "
                       "economy and cultural sector by cultural domain and "
                       "activity, 2013 + 2019-2023, long format with CV codes.",
        "source": "IBGE — Sistema de Informações e Indicadores Culturais "
                   "(SIIC), 'Informações Culturais' 2024 edition, chapter 2 "
                   "'Pesquisas estruturais em empresas'",
        "source_url": "https://www.ibge.gov.br/estatisticas/sociais/cultura.html",
        "source_files": [{"file": SOURCE_FILE.name if SOURCE_FILE else None,
                          "sha256": h}],
        "fetch_date": "2026-05-23",
        "etl_script": "etl/ibge_estruturais__siic_ch2_to_parquet.py",
        "etl_run_date": str(date.today()),
        "licence": "open data — IBGE official statistics",
        "grain": "table_id × variable × row_label × year",
        "notes": "CV = IBGE coefficient-of-variation reliability code "
                 "(A best .. E suppress; Z = value rounds to zero). value is "
                 "in the unit named in 'variable' (e.g. R$1000 for monetary "
                 "tables). row_label carries IBGE's cultural-domain hierarchy "
                 "(A.-G. domain headers + CNAE activity rows).",
    }
    out_path.with_suffix(".meta.json").write_text(
        json.dumps(meta, indent=2, ensure_ascii=False))
    print(f"  ✓ {out_path.with_suffix('.meta.json').relative_to(REPO_ROOT)}")


def maybe_push(df: pd.DataFrame, schema: str, table: str) -> None:
    """Push to MotherDuck if a valid JWT token is available.

    Skipped entirely when ATANA_ETL_SKIP_PUSH is set — used for build-only /
    sandbox runs where the Parquet is produced but the sync stays with João.
    """
    if os.environ.get("ATANA_ETL_SKIP_PUSH"):
        print(f"  · push skipped for atana.{schema}.{table} (ATANA_ETL_SKIP_PUSH)")
        return

    def _jwt(t) -> str:
        t = (t or "").strip()
        return t if (t.startswith("eyJ") and t.count(".") == 2) else ""
    token = _jwt(os.environ.get("MOTHERDUCK_TOKEN"))
    if not token:
        tf = REPO_ROOT / ".motherduck_token"
        token = _jwt(tf.read_text()) if tf.exists() else ""
    if not token:
        print(f"  · MotherDuck push skipped for atana.{schema}.{table} — no "
              f"valid token. Put a JWT in {REPO_ROOT}/.motherduck_token, re-run.")
        return
    con = duckdb.connect(f"md:atana?motherduck_token={token}")
    con.execute(f"CREATE SCHEMA IF NOT EXISTS atana.{schema}")
    con.register("df_data", df)
    con.execute(
        f"CREATE OR REPLACE TABLE atana.{schema}.{table} AS SELECT * FROM df_data")
    n = con.execute(f"SELECT COUNT(*) FROM atana.{schema}.{table}").fetchone()[0]
    print(f"  ✓ Synced atana.{schema}.{table} ({n:,} rows)")


def main() -> None:
    if not SOURCE_FILE or not SOURCE_FILE.exists():
        raise SystemExit(f"Source file not found (workspace: {WORKSPACE})")
    print(f"Reading {SOURCE_FILE.name}")
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    for table_id in SHEETS:
        if table_id not in wb.sheetnames:
            print(f"  ⚠ sheet {table_id} missing — skipped")
            continue
        df = extract_sheet(wb, table_id)
        table_name = "tab_" + table_id.replace(".", "_")
        variable = df["variable"].iloc[0] if not df.empty else "?"
        out_path = write_parquet(df, table_name)
        write_meta(out_path, table_id, variable)
        maybe_push(df, "ibge_estruturais", table_name)
    print("Done.")


if __name__ == "__main__":
    main()
